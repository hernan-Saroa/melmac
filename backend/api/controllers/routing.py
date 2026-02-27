import itertools
from api.controllers.admin import get_enterprise
from api.controllers.api import check_usage_api_enterprise, register_api_usage
from api.controllers.form import create_project_form
from api.controllers.traceability import create_traceability_service
from api.models import Address_Neighborhood, Address_UPZ, City, Service_Trazability, User_Enterprise, Massive_Errors, Massive_File, Route, Project_Enterprise, Location_Enterprise, Service_Detail, Enterprise_Service, Service_User, Service_Location, Enterprise_Process_State
from api.serializers import *
from api.permissions import IsUserAdminOrHasPermission

from django.db.models import Q
from django.http import Http404
from django.http.response import HttpResponse

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from threading import Thread
from datetime import datetime
from django.conf import settings
from django.db.utils import IntegrityError
from api.util import AddressNormalizer, RouteApi
from geopy.geocoders import GoogleV3
import pandas as pd



import json
import math
import openpyxl
import os
import pytz

TZ_INFO = pytz.timezone('America/Bogota')

NORMALIZER_INSTANCE = AddressNormalizer()

class RouteList(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)

            route_values = Route.objects.filter(
                enterprise_id=enterprise,
                state=True
            ).values(
                'id',
                'massive_file__date',
                'status'
            ).order_by(
                '-creation_date'
            )

            return Response({
                'status': True,
                'data': list(route_values)
            })
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)

    def post(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST

        if "template" in request.FILES:
            try:
                enterprise = get_enterprise(request)
                projects_ent_values = Project_Enterprise.objects.filter(enterprise_id=enterprise, state=True).values_list('identifier', flat=True)
                if len(projects_ent_values) < 1:
                    raise Exception('No hay ningun proyecto de entrega activo en el momento.')
                name_file = str(request.FILES['template'])
                if name_file.lower().endswith(('.xlsx')):
                    xlsx_file = request.FILES['template']
                    wb_obj = openpyxl.load_workbook(xlsx_file)

                    sheet = wb_obj.active

                    # print('sheet.max_row, sheet.max_column')
                    # print(sheet.max_row, sheet.max_column)

                    if sheet.max_row > 1:
                        amount_rows = sheet.max_row - 1

                        massive_val = Massive_File()
                        massive_val.enterprise_id = enterprise
                        massive_val.amount = amount_rows
                        massive_val.type = 3
                        massive_val.template = request.FILES['template']
                        massive_val.save()
                        new_thread = Thread(target=run_route, args=(massive_val, sheet, projects_ent_values,))
                        new_thread.start()

                        response['status'] = True
                        response['data'] = {
                            'id': massive_val.id,
                            'template': name_file,
                            'amount': amount_rows,
                            'success': 0,
                            'errors': 0,
                            'progressbar': 0,
                            'status': 0
                        }
                        status_response = status.HTTP_202_ACCEPTED
                        if not check_usage_api_enterprise(1, enterprise, amount_rows):
                            response['warning'] = 'Puede que este proceso falle parcialmente debido a que no tienes suficientes usos de la API de Geolocalización'
            except User_Enterprise.DoesNotExist:
                pass
            except Exception as e:
                response['detail'] = str(e)
        return Response(response, status_response)

def run_route(massive, sheet, projects_ent_values):

    data = []
    # print('massive.id')
    # print(massive.id)
    enterprise = massive.enterprise_id
    loc_file_name = '/dir.json'
    path_json_dir = settings.MEDIA_ROOT

    if not os.path.exists(path_json_dir):
        os.makedirs(path_json_dir)
    path_json_dir += loc_file_name

    try:
        content_file = ""
        with open(path_json_dir, "r") as f:
            content_file = f.read()
        loaded_addresses = json.loads(content_file)
    except:
        loaded_addresses = []
    list_addresses = { str(addr['address']) + ',' + str(addr['city']): addr for addr in loaded_addresses}

    cities_list = City.objects.filter(state=True).values_list('name', flat=True)

    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=(massive.amount + 1), max_col=13)):
        num_row = i + 2

        # columns data
        address = ''
        addressee = ''
        id_number = ''
        phone = ''
        email = ''
        city = ''
        department = ''
        weight = ''
        number_guide = ''
        observation = ''
        order = ''
        group = ''
        project = ''

        for cell in row:
            if cell.value is not None:
                if cell.column == 1:
                    address = cell.value
                elif cell.column == 2:
                    addressee = cell.value
                elif cell.column == 3:
                    id_number = cell.value
                elif cell.column == 4:
                    phone = cell.value
                elif cell.column == 5:
                    email = cell.value.strip()
                elif cell.column == 6:
                    city = cell.value.strip()
                elif cell.column == 7:
                    department = cell.value.strip()
                elif cell.column == 8:
                    weight = cell.value
                elif cell.column == 9:
                    number_guide = cell.value
                elif cell.column == 10:
                    observation = cell.value
                elif cell.column == 11:
                    order = cell.value
                elif cell.column == 12:
                    group = cell.value
                elif cell.column == 13:
                    project = str(cell.value).strip().upper()

        try:

            if address == '' or addressee == '' or city == '' or department == '' or id_number == '':
                raise Exception('Datos incompletos')

            if NORMALIZER_INSTANCE.remove_accent_mark(city).upper() not in cities_list:
                raise Exception('Ciudad no disponible')

            if project not in projects_ent_values:
                raise Exception('El proyecto no concuerda con ninguno de los identificadores ' + (','.join(projects_ent_values)))

            address_normalize = AddressNormalizer().normalize(address)
            if address_normalize:
                if address_normalize[0]['address'] == "" and address_normalize[1]['address'] == "":
                    raise Exception('La dirección no es permitida')

            search_address = address_normalize[0]['address'] if address_normalize[0]['address'] != "" else address_normalize[1]['address']
            # print(address)

            key = '{},{}'.format(address, city)
            if key not in list_addresses:

                if not check_usage_api_enterprise(1, enterprise):
                    raise Exception('Ya no tienes mas usos de la API, obten un nuevo paquete e intenta de nuevo.')
                address_info = geocode_address(search_address, city, cities_list=cities_list)

                register_api_usage(enterprise, 1, 500 if not address_info else 200)

                # time.sleep(2)
                if not address_info:
                    raise Exception('La dirección no se encontro')

                data.append({
                    'address': address,
                    'address_normalize': address_normalize,
                    'addressee': addressee,
                    'id_number': id_number,
                    'phone': phone,
                    'email': email,
                    'city': address_info['locality'],
                    'department': department,
                    'weight': weight,
                    'number_guide': number_guide,
                    'observation': observation,
                    'order': order,
                    'group': group,
                    'lat': address_info['lat'],
                    'lon': address_info['lng'],
                    'zone': address_info,
                    'project': project
                })

                loaded_addresses.append({
                    'address': address,
                    'address_normalize': address_normalize,
                    'city': city,
                    'department': department,
                    'lat': address_info['lat'],
                    'lon': address_info['lng'],
                    'zone': address_info,
                })
            else:
                temp = list_addresses[key]
                temp['addressee'] = addressee
                temp['id_number'] = id_number
                temp['phone'] = phone
                temp['email'] = email
                temp['weight'] = weight
                temp['number_guide'] = number_guide
                temp['observation'] = observation
                temp['order'] = order
                temp['group'] = group
                temp['project'] = project
                data.append(temp)
            massive.success += 1
            massive.save()
        except Exception as e:
            massive.error += 1
            data_error = (str(e) +
                ' - Dirección: ' + address +
                ', nombre destinatario: ' + addressee +
                ', ciudad: ' + city +
                ', departamento: ' + department +
                ', proyecto: ' + project
            )
            error_val = Massive_Errors()
            error_val.massive_file = massive
            error_val.row = num_row
            error_val.data = data_error
            error_val.save()

    json_data = json.dumps(data)
    # print('json_data__________:::')
    # print(json_data)

    route_val = Route()
    route_val.massive_file = massive
    route_val.status = 1
    route_val.save()

    now_time = datetime.now(tz=TZ_INFO).strftime("%d%m%Y")
    loc_file_name = str(route_val.id) + '_' + now_time + '.json'
    folder = '/' + str(massive.enterprise_id) + '/route/'
    path = settings.MEDIA_ROOT + folder

    if not os.path.exists(path):
        os.makedirs(path)
    path += loc_file_name

    with open(path, "w+") as destination:
        destination.write(json_data)

    route_val.json_path = folder + loc_file_name
    route_val.save()

    massive.status = 1
    massive.save()


    with open(path_json_dir, "w") as f:
        f.write(json.dumps(loaded_addresses))


@api_view(["POST"])
def normalizer_test(request):
    response = {
        'status': False
    }
    data = request.data
    try:
        response['status'] = True
        response['data'] = NORMALIZER_INSTANCE.normalize(data['address'])
    except:
        pass
    return Response(response)

def geocode_address(address, city, country=None, cities_list=None):
    # geolocator = Nominatim(user_agent="melmac")
    # location = geolocator.geocode(address + ", " + city + ", colombia", addressdetails=True, language="es")

    geolocator = GoogleV3(api_key="AIzaSyAeNhNt9KyTgyMCPYjVxrbqQM2kFI8ZK64")
    location = geolocator.geocode(address, components={'country':'Colombia'}, language="es")

    # print(location.raw)
    # print('{},{},{}'.format(address, city, country))
    if location:
        response = {}
        response['neighborhood'] = ''
        response['locality'] = ''
        response['sublocality'] = ''
        response['postal_code'] = ''
        response['administrative_area_level_1'] = ''
        response['country'] = ''
        response['address'] = address
        response['lat'] = location.raw['geometry']['location']['lat']
        response['lng'] = location.raw['geometry']['location']['lng']

        keys = ['neighborhood', 'locality', 'sublocality', 'postal_code', 'administrative_area_level_1', 'country']
        for key in keys:
            for part in location.raw['address_components']:
                if key in part['types']:
                    response[key] = part['long_name']
        if cities_list and NORMALIZER_INSTANCE.remove_accent_mark(response['locality'] if response['locality'] else city).upper() not in cities_list:
            # print(response)
            return None
        response['locality'] = NORMALIZER_INSTANCE.remove_accent_mark(response['locality'] if response['locality'] else city).capitalize()
        return response

    return location

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def check_status(request, pk):
    response = {
        'status': False
    }
    try:
        if 'step' not in request.GET:
            massive_val = Massive_File.objects.get(id=pk)
            progress_val = (massive_val.success + massive_val.error) / massive_val.amount * 100
            response['status'] = True
            if massive_val.status != 1:
                response['data'] = {
                    "finished": massive_val.status == 1,
                    "progress": math.ceil(progress_val)
                }
            else:
                route_val = Route.objects.get(massive_file=massive_val)
                file_path = settings.MEDIA_ROOT + str(route_val.json_path)

                content_file = ""
                with open(file_path, "r") as f:
                    content_file = f.read()

                addresses_data = json.loads(content_file)
                res = {}
                for address in addresses_data:
                    locality = address['zone']['locality']
                    if locality not in res:
                        res[locality] = {}

                    sublocality = address['zone']['sublocality']
                    if sublocality not in res[locality]:
                        res[locality][sublocality] = []

                    res[locality][sublocality].append(address)

                response['data'] = {
                    "finished": massive_val.status == 1,
                    "progress": 100,
                    "success": massive_val.success,
                    "error": massive_val.error,
                    "total": massive_val.amount,
                    "addresses": res
                }
        else:
            step = int(request.GET['step'])
            route_val = Route.objects.get(massive_file_id=pk)

            response['status'] = True
            response['data'] = {
                "finished": False,
            }
            if route_val.status in [3, 4]:
                file_path = settings.MEDIA_ROOT + str(route_val.json_path)

                content_file = ""
                with open(file_path, "r") as f:
                    content_file = f.read()

                addresses_data = json.loads(content_file)

                services = Service_Detail.objects.filter(service__state=True, service__route_file=route_val).select_related('initial_position').order_by('service_id').values('service_id', 'id', 'initial_position__address', 'initial_position__lat', 'initial_position__lon', 'api_route')

                ini_point = {'address': services[0]['initial_position__address'], 'address_normalized': services[0]['initial_position__address'], 'lat': services[0]['initial_position__lat'], 'lon': services[0]['initial_position__lon'], 'order': 0, 'is_ent_location': True}

                get_attr = lambda point: point['service_id'] if 'service_id' in point else 0
                services_list = {k: {'points':[ini_point]+list(g)} for k, g in itertools.groupby(sorted(addresses_data, key=get_attr), get_attr)}
                for service in services:
                    key = service['service_id']
                    if service['api_route']:
                        file_path = settings.MEDIA_ROOT +'/' + service['api_route']
                        content_file = ""
                        with open(file_path, "r") as f:
                            content_file = f.read()

                        path_data = json.loads(content_file)
                        services_list[key]['path'] = path_data

                        services_list[key]['points'].sort(key=lambda x: x['order'])
                try:
                    del services_list[0]
                except:
                    pass
                response['data'] = {
                    "finished": (route_val.status == 3 and step == 2) or (route_val.status == 4 and step == 3),
                    "routes": services_list
                }


    except (Route.DoesNotExist, Massive_File.DoesNotExist):
        response['message'] = "El elemento no existe"

    return Response(response)

class ProjectList(APIView):
    permission_classes = [IsUserAdminOrHasPermission]

    def get(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            project_values = Project_Enterprise.objects.filter(
                enterprise_id=enterprise
            ).values(
                'id',
                'identifier',
                'name',
                'description',
                'created_at',
                'modified_at',
                'state'
            ).order_by(
                '-created_at'
            )

            response = {
                'status': True,
                'data': list(project_values)
            }
            status_response = status.HTTP_200_OK
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)

    def post(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            data = request.data

            project_forms_val = Project_Enterprise.objects.filter(form__isnull=False, enterprise_id=enterprise)
            if len(project_forms_val) < 1:
                project_forms_val = create_project_form(request.user)
            else:
                project_forms_val = project_forms_val[0].form_id

            if project_forms_val is None:
                raise Exception("No se logro vincular el documento de entrega al proyecto")

            projects_not_assigned = Project_Enterprise.objects.filter(form__isnull=True, enterprise_id=enterprise)
            if len(projects_not_assigned) > 0:
                projects_not_assigned.update(form_id=project_forms_val)

            project_val = Project_Enterprise()
            project_val.enterprise_id = enterprise
            project_val.name = data['name']
            project_val.identifier = data['identifier'].strip().upper()
            project_val.description = data['description']
            project_val.form_id = project_forms_val


            project_val.save()

            response['status'] = True
            response['data'] = {
                'id': project_val.id,
                'name': data['name'],
                'identifier': data['identifier'],
                'description': data['description'],
                'state': True,
            }

            status_response = status.HTTP_200_OK

        except User_Enterprise.DoesNotExist:
            pass
        except IntegrityError:
            response['detail'] = "El valor de la columna Identificador ya se encuentra registrada con otro proyecto y este valor debe ser unico."
        except Exception as e:
            print(e)
            response['detail'] = str(e)
        return Response(response, status_response)

class ProjectDetail(APIView):
    permission_classes = [IsUserAdminOrHasPermission]

    def get_object(self, pk, enterprise_id):
        try:
            return Project_Enterprise.objects.get(pk=pk, enterprise_id=enterprise_id)
        except Project_Enterprise.DoesNotExist:
            raise Http404

    def put(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            data = request.data
            project_val = self.get_object(pk, enterprise)

            project_val.name = data['name']
            project_val.description = data['description']
            project_val.state = data['state'] == 'true'

            project_val.modified_at = datetime.now()
            project_val.save()

            response = {"status": True}
            status_response = status.HTTP_200_OK
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)

    def delete(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            data = request.data
            project_val = self.get_object(pk, enterprise)

            project_val.modified_at = datetime.now()
            project_val.state = False
            project_val.save()

            response['status'] = True

            status_response = status.HTTP_200_OK

        except User_Enterprise.DoesNotExist:
            pass
        except Exception as e:
            response['detail'] = e.message
        return Response(response, status_response)

class LocationList(APIView):
    permission_classes = [IsUserAdminOrHasPermission]

    def get(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            project_values = Location_Enterprise.objects.filter(
                enterprise_id=enterprise,
                state=True
            ).values(
                'id',
                'address',
                'city',
                'country',
                'name',
                'lat',
                'lon',
                'created_at',
                'modified_at',
                'is_default',
                'state'
            ).order_by(
                '-created_at'
            )

            response = {
                'status': True,
                'data': list(project_values)
            }
            status_response = status.HTTP_200_OK
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)

    def post(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            data = request.data
            cities = City.objects.filter(state=True).select_related('country').values('id', 'name', 'country_id', 'country__name')
            cities_list = {city['id']: city for city in cities}
            address = AddressNormalizer().normalize(data['address'])
            if address:
                if address[0]['address'] == "" and address[1]['address'] == "":
                    raise Exception('La dirección no es permitida')

            address_normalized = address[0]['address'] if address[0]['address'] != "" else address[1]['address']
            city_name = cities_list[data['city']]['name']
            country_name = cities_list[data['city']]['country__name']

            address_info = geocode_address(address_normalized, city_name, country_name)

            if not address_info:
                raise Exception('La dirección no se encontro')

            # print(address_info.raw)

            location_val = Location_Enterprise()
            location_val.enterprise_id = enterprise
            location_val.name = data['name']
            location_val.address = address_normalized
            location_val.city_id = data['city']
            location_val.country_id = cities_list[data['city']]['country_id']
            location_val.lat = address_info['lat']
            location_val.lon = address_info['lng']
            location_val.is_default = data['is_default']

            location_val.save()

            if location_val.is_default:
                Location_Enterprise.objects.filter(enterprise_id=enterprise).exclude(id=location_val.id).update(is_default=False)

            response['status'] = True
            response['data'] = {
                'id': location_val.id,
                'name': data['name'],
                'address': address_normalized,
                'city': location_val.city_id,
                'country': location_val.country_id,
                'lat': location_val.lat,
                'lon': location_val.lon,
                'is_default': location_val.is_default
            }

            status_response = status.HTTP_200_OK

        except User_Enterprise.DoesNotExist:
            pass
        except Exception as e:
            print(e)
            response['detail'] = str(e)
        return Response(response, status_response)

class LocationDetail(APIView):
    permission_classes = [IsUserAdminOrHasPermission]

    def get_object(self, pk, enterprise_id):
        try:
            return Location_Enterprise.objects.get(pk=pk, enterprise_id=enterprise_id)
        except Location_Enterprise.DoesNotExist:
            raise Http404

    def put(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            cities = City.objects.filter(state=True).select_related('country').values('id', 'name', 'country_id', 'country__name')
            cities_list = {city['id']: city for city in cities}
            data = request.data
            location_val = self.get_object(pk, enterprise)
            address_update = False

            if data['address'] != location_val.address or data['city'] != location_val.city or data['city'] != location_val.country:
                address_update = True
                address = AddressNormalizer().normalize(data['address'])

                city_name = cities_list[data['city']]['name']
                country_name = cities_list[data['city']]['country__name']
                if address:
                    if address[0]['address'] == "" and address[1]['address'] == "":
                        raise Exception('La dirección no es permitida')
                address_normalized = address[0]['address'] if address[0]['address'] != "" else address[1]['address']
                address_info = geocode_address(address_normalized, city_name, country_name)

                if not address_info:
                    raise Exception('La dirección no se encontro')

            location_val.name = data['name']
            if address_update:
                location_val.address = address_normalized
                location_val.city_id = data['city']
                location_val.country_id = cities_list[data['city']]['country_id']
                location_val.lat = address_info['lat']
                location_val.lon = address_info['lng']
                location_val.is_default = data['is_default']

            location_val.save()


            if location_val.is_default:
                Location_Enterprise.objects.filter(enterprise_id=enterprise).exclude(id=location_val.id).update(is_default=False)

            response['status'] = True
            response['data'] = {
                'id': location_val.id,
                'name': data['name'],
                'address': location_val.address,
                'city': location_val.city_id,
                'country': location_val.country_id,
                'lat': location_val.lat,
                'lon': location_val.lon,
                'is_default': location_val.is_default
            }

            status_response = status.HTTP_200_OK
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)

    def delete(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            location_val = self.get_object(pk, enterprise)

            location_val.state = False

            location_val.save()

            response['status'] = True
            status_response = status.HTTP_200_OK

        except User_Enterprise.DoesNotExist:
            pass
        except Exception as e:
            response['detail'] = e.message
        return Response(response, status_response)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_address_processes(request):
    response = {"status": False}
    status_response = status.HTTP_400_BAD_REQUEST
    try:
        enterprise = get_enterprise(request)
        # Todas las cargas de direcciones
        all_address_uploads = Massive_File.objects.filter(enterprise_id=enterprise, type=3)
        # Todos los json generados con direcciones
        completed_uploads = Route.objects.filter(massive_file__enterprise_id=enterprise, state=True)

        completed_ids = [complete.massive_file_id for complete in list(completed_uploads)]

        response['data'] = {}
        response['data']['complete'] = list(completed_uploads.exclude(status__in=[1,2,3]).select_related('massive_file').order_by('-massive_file__date').values('massive_file_id', 'massive_file__amount', 'massive_file__success', 'massive_file__error', 'massive_file__date', 'id'))
        response['data']['incomplete'] = list(completed_uploads.exclude(status=4).select_related('massive_file').order_by('-massive_file__date').values('massive_file_id', 'massive_file__amount', 'massive_file__success', 'massive_file__error', 'massive_file__date', 'id', 'status'))
        temp = []
        for massive_file in list(all_address_uploads.exclude(id__in=completed_ids).values()):
            temp.append({
                'massive_file_id': massive_file['id'],
                'massive_file__amount': massive_file['amount'],
                'massive_file__success': massive_file['success'],
                'massive_file__error': massive_file['error'],
                'massive_file__date': massive_file['date'],
                'id': None,
                'status': 0,
            })
        response['data']['incomplete'] += temp
        response['status'] = True
        status_response = status.HTTP_200_OK

    except User_Enterprise.DoesNotExist:
        pass
    except Exception as e:
        response['detail'] = e.message
    return Response(response, status_response)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_address_process(request, pk):
    response = {"status": False}
    status_response = status.HTTP_400_BAD_REQUEST
    try:
        step = 0
        res = {}
        enterprise = get_enterprise(request)
        try:
            route_val = Route.objects.get(massive_file_id=pk, state=True)
            step = route_val.status

            file_path = settings.MEDIA_ROOT + str(route_val.json_path)

            content_file = ""
            with open(file_path, "r") as f:
                content_file = f.read()

            addresses_data = json.loads(content_file)

            # Json Generado
            if step == 1:
                for address in addresses_data:
                    city = address['zone']['locality']
                    if city not in res:
                        res[city] = {}

                    locality = address['zone']['sublocality']
                    if locality not in res[city]:
                        res[city][locality] = []

                    res[city][locality].append(address)

                massive_val = route_val.massive_file
                response['data'] = {
                    "finished": massive_val.status == 1,
                    "progress": 100,
                    "success": massive_val.success,
                    "error": massive_val.error,
                    "total": massive_val.amount,
                    "addresses": res
                }
            elif step in [2,3]:
                # Generando Rutas
                initial_point = Service_Detail.objects.filter(service__state=True, service__route_file_id=route_val.id).select_related('initial_position').values('initial_position__address', 'initial_position__lat', 'initial_position__lon')
                if initial_point:
                    initial_point = initial_point[0]
                    ini_point = {'address': initial_point['initial_position__address'], 'address_normalized': initial_point['initial_position__address'], 'lat': initial_point['initial_position__lat'], 'lon': initial_point['initial_position__lon'], 'order': 0, 'is_ent_location': True}

                    get_attr = lambda point: point['service_id'] if 'service_id' in point else 0
                    new_list = {k: {'points':[ini_point]+list(g)} for k, g in itertools.groupby(sorted(addresses_data, key=get_attr), get_attr)}
                    try:
                        del new_list[0]
                    except:
                        pass
                    response['data'] = {'routes': new_list}
                    response['detail'] = list(Service_Detail.objects.filter(service__route_file_id=route_val.id, service__state=True).values())

        except Route.DoesNotExist:
            massive_val = Massive_File.objects.get(id=pk, enterprise_id=enterprise, type=3)
            response['data'] = {
                "finished": massive_val.status == 1,
                "progress": 100,
                "success": massive_val.success,
                "error": massive_val.error,
                "total": massive_val.amount,
                "addresses": res
            }

        response['step'] = step
        response['status'] = True
        status_response = status.HTTP_200_OK

    except User_Enterprise.DoesNotExist:
        pass
    except Massive_File.DoesNotExist:
        response['detail'] = "No existe este registro"
    # except Exception as e:
    #     response['detail'] = str(e)
    return Response(response, status_response)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def list_services(request):
    response = {}
    response['status'] = False
    user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
    service_list_free = []
    service_list_user = []
    try:
        # Servicios del usuario
        service_user_vals = Service_User.objects.filter(user=user_val).select_related('service').values_list('service__service_id', flat=True)

        # Obtiene los del usuario y los servicios libres
        service_values = Service_Detail.objects.filter(
            Q(process_state__order__in=[1]) | Q(service_id__in=service_user_vals),
            service__enterprise_id=user_val.enterprise_id,
            service__state=True,
            process_state__process_id=1,
        ).select_related(
            'service'
        ).values(
            'service_id',
            'process_state_id',
            'name',
            'description',
            'initial_addres'
        )
        for service_val in service_values:
            # Opcion 1 servicio libre
            if service_val['process_state_id'] == 1:
                service_list_free.append({
                    'id': service_val['service_id'],
                    'name': service_val['name'],
                    'description': service_val['description'],
                    'address' : service_val['initial_addres']
                })
            # Opcion 3 servicio del usuario
            elif service_val['process_state_id'] == 3:
                service_list_user.append({
                    'id': service_val['service_id'],
                    'name': service_val['name'],
                    'process_state_id' : service_val['process_state_id'],
                    'description': service_val['description'],
                    'address' : service_val['initial_addres']
                })
        response['status'] = True
        response['state_1'] = service_list_free
        response['state_3'] = service_list_user
    except Enterprise_Service.DoesNotExist:
        response['message'] = 'No hay servicios para mostrar'
    except Exception as error:
        print(error)
        response['message'] = 'Algo salio mal, verifica la información enviada ' + str(error)
    return Response(response)

def zonify_location(location, address, aux):
    neighborhood = NORMALIZER_INSTANCE.remove_accent_mark(address['zone']['neighborhood']).upper()
    postal_code = address['zone']['postal_code']
    city = NORMALIZER_INSTANCE.remove_accent_mark(address['zone']['administrative_area_level_1']).upper()
    lat = address['zone']['lat']
    lon = address['zone']['lng']

    if neighborhood in aux['neighborhood']:
        location.neighborhood_id = aux['neighborhood'][neighborhood]['id']
        location.upz_id = aux['neighborhood'][neighborhood]['upz_id']
        location.locality_id = aux['upz'][location.upz_id]['locality_id']

    if postal_code!="" and postal_code.isnumeric():
        location.postal_code = int(postal_code)

    if city in aux['city']:
        location.city_id = aux['city'][city]['id']
        location.country_id = aux['city'][city]['country_id']
    else:
        location.city_id = 1
        location.country_id = 1

    location.latitude = lat
    location.longitude = lon

    return location


def generate_routes(initial_location, num_messagers, user, route_val, aux):

    file_path = settings.MEDIA_ROOT + str(route_val.json_path)

    content_file = ""
    with open(file_path, "r") as f:
        content_file = f.read()

    addresses_data = json.loads(content_file)

    process_state = Enterprise_Process_State.objects.get(process__name="ENRUTAMIENTO", order=3)

    process_state_location = Enterprise_Process_State.objects.get(process__name="MENSAJERIA", order=1)

    # print(range(num_messagers))
    services_route = []
    for i in range(num_messagers):
        # Service
        temp = Enterprise_Service()
        temp.created_by = user
        temp.enterprise_id = user.enterprise_id
        temp.route_file = route_val
        temp.save()

        # Service Detail
        temp2 = Service_Detail()
        temp2.service = temp
        temp2.initial_position = initial_location
        temp2.process_state = process_state
        temp2.initial_addres = initial_location.address
        temp2.location_quantity = 0
        temp2.total_distance = 0
        services_route.append({'service': temp, 'quantity': 0, 'detail': temp2, 'points': []})

    # print(len(services_route))
    loaded_addresses = route_val.massive_file.success

    service_index = 0
    services_assigned = 0
    # Location
    num_messagers_temp = num_messagers
    for i, address in enumerate(addresses_data):
        try:
        # if True:
            temp = Service_Location()
            temp.service = services_route[service_index]['service']
            temp.route_file = route_val
            temp.order = 0
            temp.address = address['address']
            temp.address_normalized = address['address_normalize'][0]['address']
            temp.comment = address['observation']
            temp.comment_address = address['address_normalize'][0]['comment']
            temp.user_name = address['addressee']
            temp.user_phone_number = address['phone']
            temp.user_email = address['email']
            temp.guide_number = address['number_guide']
            temp = zonify_location(temp, address, aux)
            temp.process_state = process_state_location
            temp.project_id = aux['project'][address['project']]
            temp.save()
            ent_text = '0'*(6-len(str(user.enterprise_id))) + str(user.enterprise_id)
            id_text = '0'*(10-len(str(temp.id))) + str(temp.id)
            temp.guide_number_internal = 'ME{}{}'.format(ent_text,id_text)
            temp.save()


            services_route[service_index]['points'].append(temp)
            services_route[service_index]['quantity'] += 1
            services_assigned += 1

            # print(loaded_addresses, num_messagers_temp, loaded_addresses/num_messagers)

            if num_messagers_temp > 1 and services_route[service_index]['quantity'] >= (loaded_addresses/num_messagers_temp):
                loaded_addresses -= services_route[service_index]['quantity']
                service_index += 1
                num_messagers_temp = (num_messagers_temp - 1)

            addresses_data[i]['service_id'] = temp.service.id
            addresses_data[i]['location_id'] = temp.id
        except :
            loaded_addresses -= 1
            pass

    route_val.status = 2
    route_val.save()

    optimize_routes(services_route, addresses_data, file_path, enterprise=user.enterprise_id)


def optimize_routes(routes, addresses_data, file_path, enterprise, edit=False, edit_extra=None):
    # print(routes)
    for route in routes:
        path_info = {
            'path':[],
            'time': 0,
            'distance': 0
        }
        points = route['points']
        initial_point = route['detail'].initial_position
        initial_point = [initial_point.lat, initial_point.lon]
        option_points_og = [[point.latitude, point.longitude] for point in points]
        temp_point = initial_point
        initial = initial_point
        coords = [initial] + option_points_og
        # print(coords)
        order_list = RouteApi().driver(coords)
        order_point_list = [order_list, []]

        for index in order_list[1:]:
            order_point_list[1].append(coords[index-2])

        i = 0
        while True:
            if i > len(order_point_list[1]):
                break
            temp_points = order_point_list[1][i:i+25]
            api_route = RouteApi().get_response(initial_point, temp_points[-1], temp_points[:-1])
            status = 200 if len(api_route) else 500
            register_api_usage(enterprise, 2, status, 'Rutas generadas' if len(api_route) else api_route['status'])

            partial_distance = 0
            partial_time = 0
            if api_route:
                if 'path' not in route:
                    route['path'] = []

                for leg in api_route[0]['legs']:
                    partial_distance += leg['distance']['value']
                    partial_time += leg['duration']['value']
                    for step in leg['steps']:
                        path_info['path'].append(step['start_location'])
                        path_info['path'].append(step['end_location'])
            path_info['distance'] = partial_distance
            path_info['time'] = partial_time

            i += 25
            initial_point = temp_points[-1]

        for i, coord in enumerate(order_list[1:]):
            temp_point = route['points'][coord-2]
            temp_point.order = i+1
            temp_point.save()
            file_index = [index_i for index_i, address in enumerate(addresses_data) if 'location_id' in address and address['location_id'] == temp_point.id][0]
            addresses_data[file_index]['order'] = i+1

        route['detail'].total_distance = path_info['distance']
        route['detail'].total_time = path_info['time']
        route['path'] = path_info
        file_name = str(enterprise) + '/' + str(route['detail'].service_id) + '/path.json'
        file_route_path = os.path.join(settings.MEDIA_ROOT, str(enterprise) + '/' + str(route['detail'].service_id))
        if not os.path.exists(file_route_path):
            os.makedirs(file_route_path)
        file_route_path = os.path.join(settings.MEDIA_ROOT, file_name)
        with open(file_route_path, "w") as f:
            f.write(json.dumps(path_info))
        route['detail'].api_route = file_name
        route['detail'].location_quantity = len(route['points'])
        route['detail'].save()

    if edit:
        for i, address in enumerate(addresses_data):
            if address['location_id'] == edit_extra['location']:
                # print(edit_extra, '{} => {}'.format(addresses_data[file_index]['service_id'], edit_extra['service']))
                addresses_data[i]['service_id'] = edit_extra['service']
                break


    with open(file_path, "w") as f:
        f.write(json.dumps(addresses_data))

    if not edit:
        route = routes[0]['service'].route_file
        route.status = 3
        route.save()
    else:
        return addresses_data


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_route_services(request):
    response = {"status": False}
    status_response = status.HTTP_400_BAD_REQUEST
    try:
        data = request.data
        pk = data['id']
        num_messagers = data['num_messagers']
        user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
        location = Location_Enterprise.objects.get(id=data['location'])

        aux = {}

        temp = Address_Neighborhood.objects.filter(state=True).values('id', 'upz_id', 'name')
        temp = {neighbor['name']: neighbor for neighbor in list(temp)}

        aux['neighborhood'] = temp

        temp = Address_UPZ.objects.filter(state=True).values('id', 'name', 'locality_id')
        temp = {upz['id']: upz for upz in list(temp)}

        aux['upz'] = temp

        temp = City.objects.filter(state=True).values('id', 'name', 'country_id')
        temp = {city['name']: city for city in list(temp)}

        aux['city'] = temp

        temp = Project_Enterprise.objects.filter(enterprise_id=user_val.enterprise_id, state=True).values('id', 'identifier')
        temp = {project['identifier']: project['id'] for project in list(temp)}

        aux['project'] = temp

        try:
            route_val = Route.objects.get(massive_file_id=pk, status=1, state=True)
            new_thread = Thread(target=generate_routes, args=(location, num_messagers, user_val, route_val, aux,))
            new_thread.start()

            # generate_routes(location, num_messagers, user_val, route_val, aux)

            response['status'] = True
            response['internal_step'] = 2
            status_response = status.HTTP_200_OK

            if not check_usage_api_enterprise(2, user_val.enterprise_id, num_messagers):
                response['warning'] = 'Puede que este proceso generé costos adicionales por usos de la API de Generación de Rutas'
        except Route.DoesNotExist:
            response['status'] = False
            status_response = status.HTTP_200_OK
            route_val = Route.objects.get(massive_file_id=pk, state=True)
            response['internal_step'] = route_val.status

    except User_Enterprise.DoesNotExist:
        pass
    # except Exception as e:
    #     response['detail'] = str(e)
    return Response(response, status_response)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def route_services(request):
    response = {}
    response['status'] = False
    try:
        data = request.data
        route_service_val = Service_Location.objects.filter(service_id = data['service_id'], state = True, process_state_id = 3).select_related('project').order_by('-order').values('id', 'comment', 'user_name', 'address_normalized', 'guide_number', 'user_phone_number','order', 'project__form')
        if len(route_service_val) != 0:
            for route in route_service_val:
                print("")
            response['status'] = True
            response['route'] = route
        else:
            response['status'] = False
            response['message'] = "No hay rutas para mostrar"
    except Service_Location.DoesNotExist:
        response['message'] = 'No hay rutas para mostrar'
    except Exception as error:
        print(error)
        response['message'] = 'Algo salio mal, verifica la información enviada ' + str(error)
    return Response(response)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_location_to_route(request):
    status_response = status.HTTP_400_BAD_REQUEST
    response = {}
    response['status'] = False
    try:
        data = request.data
        user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
        location = Service_Location.objects.get(id=data['location'], state=True)
        service_initial = location.service
        service_final = Enterprise_Service.objects.get(id=data['final'], state=True)
        location.service_id = data['final']
        location.save()

        route_val = service_final.route_file

        file_path = settings.MEDIA_ROOT + str(route_val.json_path)

        content_file = ""
        with open(file_path, "r") as f:
            content_file = f.read()

        addresses_data = json.loads(content_file)

        routes_val = []

        for service in [service_initial, service_final]:
            # print(service)
            temp = {}
            temp['service'] = service
            temp['detail'] = Service_Detail.objects.get(service_id=service.id, service__state=True)
            temp['detail'].total_distance = 0
            temp['points'] = list(Service_Location.objects.filter(service_id=service.id, state=True))
            routes_val.append(temp)

        initial_position = routes_val[0]['detail'].initial_position

        points_organized = optimize_routes(routes_val, addresses_data, file_path, enterprise=user_val.enterprise_id, edit=True, edit_extra={'location': location.id, 'service': service_final.id})

        services_list = {}
        for point in list(points_organized):
            key = point['service_id'] if 'service_id' in point else 0
            if key > 0:
                if key not in services_list:
                    services_list[key] = {
                        'points': [{
                            'address': initial_position.address,
                            'lat': initial_position.lat,
                            'lon': initial_position.lon,
                            'order': 0,
                            }]
                    }
                services_list[key]['points'].append(point)
                services_list[key]['points'].sort(key=lambda x: x['order'])

        for service in routes_val:
            file_path = settings.MEDIA_ROOT +'/'+ str(service['detail'].api_route)
            path_data = []
            # try:
            content_file = ""
            with open(file_path, "r") as f:
                content_file = f.read()
            path_data = json.loads(content_file)
            services_list[service['service'].id]['path'] = path_data
            # except Exception as e:
            #     print(e)


        if not check_usage_api_enterprise(2, user_val.enterprise_id):
            response['warning'] = 'Puede que este proceso generó costos adicionales por usos de la API de Generación de Rutas'

        response['status'] = True
        response['data'] = services_list
        status_response = status.HTTP_200_OK

    except User_Enterprise.DoesNotExist:
        pass
    # except Exception as e:
    #     response['detail'] = str(e)

    return Response(response, status_response)


#process_id = 0: comnfirmando servicio por parte del mensajero y cambio de estado de las rutas a asignado a mensajero
#process_id = 1:
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_state(request):
    response = {}
    response['status'] = False
    try:
        trace_data = None
        data = request.data
        # print(data)
        if data['process_id'] == 0:
            Service_Detail.objects.filter(service_id = data['service_id']).update(process_state_id = 3)
            Service_Location.objects.filter(service_id = data['service_id']).update(process_state_id = 3)
            response['status'] = True
            response['message'] = "Estado del servicio y de las rutas actualizado a Asignado a mensajero"
        elif data['process_id'] == 1:
            Service_Detail.objects.filter(service_id = data['service_id']).update(process_state_id = 2)
            Service_Location.objects.filter(id = data['id']).update(process_state_id = 2)
            response['status'] = True
            response['message'] = "Estado de la ruta actualizado a En ruta"
            # llamado de cambio de estado Trazabilidad
            trace_data = [{
                'process': 1,
                'group': 2,
                'reference': data['id'],
                'description': '',
                'latitude': data['latitude'],
                'longitude': data['longitude'],
                'process_state': 2,
            }]
        elif data['process_id'] == 2:
            Service_Location.objects.filter(id = data['id']).update(process_state_id = 6)
            response['status'] = True
            response['message'] = "Estado de la ruta actualizado a Llegada a punto de entrega"
            trace_data = [{
                'process': 1,
                'group': 2,
                'reference': data['id'],
                'description': '',
                'latitude': data['latitude'],
                'longitude': data['longitude'],
                'process_state': 6,
            }]
        elif data['process_id'] == 3:
            Service_Location.objects.filter(id = data['id']).update(process_state_id = 7)
            response['status'] = True
            response['message'] = "Estado de la ruta actualizado a En proceso de Captura"
            trace_data = [{
                'process': 1,
                'group': 2,
                'reference': data['id'],
                'description': '',
                'latitude': data['latitude'],
                'longitude': data['longitude'],
                'process_state': 7,
            }]
        elif data['process_id'] == 4:
            Service_Location.objects.filter(id = data['id']).update(process_state_id = 8)
            response['status'] = True
            response['message'] = "Estado de la ruta actualizado a Entrega Finalizada"
        elif data['process_id'] == 5:
            Service_Detail.objects.filter(service_id = data['service_id']).update(process_state_id = 8)
            response['status'] = True
            response['message'] = "Estado de la ruta actualizado a Entrega Finalizada"
            # llamado de cambio de estado Trazabilidad
            trace_data = [{
                'process': 1,
                'group': 2,
                'reference': data['id'],
                'description': '',
                'latitude': data['latitude'],
                'longitude': data['longitude'],
                'process_state': 8,
            }]
        elif data['process_id'] == 6:
            Service_Location.objects.filter(id = data['id']).update(process_state_id = 13)
            response['status'] = True
            response['message'] = "Estado de la ruta actualizado a Servicio Cancelado"
        elif data['process_id'] == 7:
            Service_Detail.objects.filter(service_id = data['service_id']).update(process_state_id=8)
            Service_Location.objects.filter(service_id = data['service_id']).update(process_state_id=8)
            Service_User.objects.filter(service__service_id = data['service_id']).update(process_state_id=8)
            response['status'] = True
            response['message'] = "Estado del servicio actualizado"
            service_location_values = Service_Location.objects.filter(service_id = data['service_id'])
            # llamado de cambio de estado Trazabilidad
            for service_location_val in service_location_values:
                trace = {
                    'process': 1,
                    'group': 2,
                    'reference': service_location_val.id,
                    'description': '',
                    'latitude': service_location_val.latitude,
                    'longitude': service_location_val.longitude,
                    'process_state': 8,
                }
                create_traceability_service(trace)
        else:
            response['status'] = False
            response['message'] = "No hay rutas para mostrar"

        if trace_data:
            for trace in trace_data:
                create_traceability_service(trace)
    except Service_Location.DoesNotExist:
        response['message'] = 'No hay rutas para mostrar'
    except Exception as error:
        print(error)
        response['message'] = 'Algo salio mal, verifica la información enviada ' + str(error)
    return Response(response)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def assign_messagers(request):
    status_response = status.HTTP_400_BAD_REQUEST
    response = {}
    response['status'] = False
    try:
        data = request.data
        user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
        pk = data['id']
        route_val = Route.objects.get(massive_file_id=pk, status=3, state=True)

        file_path = settings.MEDIA_ROOT + str(route_val.json_path)

        content_file = ""
        with open(file_path, "r") as f:
            content_file = f.read()

        addresses_data = json.loads(content_file)

        route_points = Service_Location.objects.filter(state=True, service_id__in=data['assign'].keys()).values('id', 'service_id')

        process_state_new = Enterprise_Process_State.objects.get(process__name="MENSAJERIA", order=1)
        process = process_state_new.process
        process_state_assigned = Enterprise_Process_State.objects.get(process__name="MENSAJERIA", order=3)

        if route_points:
            get_attr = lambda point: point['service_id'] if 'service_id' in point else 0
            locations_list = {k: list(g) for k, g in itertools.groupby(sorted(route_points, key=get_attr), get_attr)}

            try:
                del locations_list[0]
            except:
                pass

            objs = []
            trace_objs = []
            now = datetime.now(tz=TZ_INFO)
            now_date = now.strftime("%d/%m/%Y")
            now_time = now.strftime("%H:%M")

            for key, assigned in data['assign'].items():
                only_create = False

                for location in locations_list[int(key)]:
                    if assigned:
                        objs.append(Service_User(service_id=location['id'], distance=0, user_id=assigned, process_state=process_state_assigned, distance_api="", time_api=""))
                    else:
                        trace_objs.append(Service_Trazability(process=process, group=2, reference=location['id'], date_trace=now_date, hour_trace=now_time, latitude=0, longitude=0, process_state=process_state_new))
                        only_create = True

                if only_create:
                    Service_Detail.objects.filter(service_id=key).update(process_state=process_state_new)
                else:
                    Service_Detail.objects.filter(service_id=key).update(process_state=process_state_assigned)


                for i, address in enumerate(addresses_data):
                    if 'service_id' in address and address['service_id'] == key:
                        addresses_data[i]['user_id'] = assigned

            objs = Service_User.objects.bulk_create(objs)

            for obj in objs:
                trace_objs.append(Service_Trazability(process=process, group=2, reference=obj.service_id, date_trace=now_date, hour_trace=now_time, latitude=0, longitude=0, process_state=process_state_new))
                trace_objs.append(Service_Trazability(process=process, group=2, reference=obj.service_id, date_trace=now_date, hour_trace=now_time, latitude=0, longitude=0, process_state=process_state_assigned))

            Service_Trazability.objects.bulk_create(trace_objs)

            with open(file_path, "w") as f:
                f.write(json.dumps(addresses_data))

            route_val.status = 4
            route_val.save()

            status_response = status.HTTP_200_OK
            response['status'] = True
    except User_Enterprise.DoesNotExist:
        pass
    except Exception as error:
        print(error)
        response['message'] = 'Algo salio mal, verifica la información enviada ' + str(error)

    return Response(response, status_response)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def services_monitor(request):
    response = {}
    status_response = status.HTTP_400_BAD_REQUEST
    response['status'] = False
    try:
        user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)

        service_list = []

        if user_val.role_id == 2 or user_val.role_enterprise.is_admin:
            service_enterprise_val = Service_Detail.objects.filter(service__enterprise_id=user_val.enterprise_id, service__state=True).select_related('process_state').order_by('-service_id').values('service_id','location_quantity', 'process_state__name', 'name', 'description')
            service_list = list(service_enterprise_val)
        else:
            role_values = Permit_Role.objects.filter(
                role_enterprise__enterprise_id=user_val.enterprise_id,
                role_enterprise__state=True,
                permit__permit_type_id=43,
                state=True
            ).select_related(
                'role_enterprise',
                'permit',
            ).values_list(
                'role_enterprise_id', flat=True
            )
            if user_val.role_enterprise_id in role_values:
                service_user_val = Service_User.objects.filter(user_id=user_val.id).select_related('service').values_list('service__service_id', flat=True)
                service_enterprise_val = Service_Detail.objects.filter(service_id__in=service_user_val, service__state=True).select_related('process_state').order_by('-service_id').values('service_id','location_quantity', 'process_state__name', 'name', 'description')
                service_list = list(service_enterprise_val)

        response['status'] = True
        status_response = status.HTTP_200_OK
        response['data'] = service_list
    except User_Enterprise.DoesNotExist:
        pass
    except Exception as error:
        response['message'] = 'Algo salio mal, contacta al administrador para solucionarlo ' + str(error)

    return Response(response, status_response)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def services_monitor_detail(request, pk):
    response = {}
    status_response = status.HTTP_400_BAD_REQUEST
    response['status'] = False
    try:
        user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
        service_list = []
        path_data = {}

        service_detail_val = Service_Detail.objects.get(service_id=pk, service__state=True)
        service_user_val = Service_User.objects.filter(
            service__service_id=pk, state=True
        ).select_related(
            'service',
            'user'
        ).order_by('service__order').values(
            'service_id',
            'service__address_normalized',
            'service__comment',
            'service__latitude',
            'service__longitude',
            'service__user_name',
            'service__user_email',
            'service__user_phone_number',
            'service__user_id_number',
            'service__guide_number',
            'service__process_state__name',
            'service__answer_form_id',
            'user__first_name',
            'user__first_last_name',
            'user__phone',
            'user__email',
            'service__order'
        )
        # print(service_user_val.query)
        if not service_user_val:
            service_user_val = Service_Location.objects.filter(service_id=pk, state=True).select_related('process_state').order_by('order').values('id', 'address_normalized', 'comment', 'latitude', 'longitude', 'answer_form_id', 'user_name', 'user_email', 'user_phone_number', 'user_id_number', 'guide_number', 'order', 'process_state__name', 'service_id')
            service_user_val = [
                {
                    'service_id': location['id'],
                    'service__address_normalized': location['address_normalized'],
                    'service__comment': location['comment'],
                    'service__latitude': location['latitude'],
                    'service__longitude': location['longitude'],
                    'service__user_name': location['user_name'],
                    'service__user_email': location['user_email'],
                    'service__user_phone_number': location['user_phone_number'],
                    'service__user_id_number': location['user_id_number'],
                    'service__guide_number': location['guide_number'],
                    'service__process_state__name': location['process_state__name'],
                    'service__order': location['order'],
                    'service__answer_form_id': location['answer_form_id'],
                    'user__first_name': 'N/A',
                    'user__first_last_name': '',
                    'user__phone': '',
                    'user__email': ''
                }
                for location in service_user_val
            ]
            file_path = settings.MEDIA_ROOT +'/'+ str(service_detail_val.api_route)
            # try:
            content_file = ""
            with open(file_path, "r") as f:
                content_file = f.read()
            path_data = json.loads(content_file)

        response['status'] = True
        status_response = status.HTTP_200_OK
        response['data'] = {
            'details': {
                'initial_position__address': service_detail_val.initial_position.address,
                'lat': service_detail_val.initial_position.lat,
                'lon': service_detail_val.initial_position.lon,
                'total_distance': service_detail_val.total_distance,
                'total_duration': path_data['time'] if 'time' in path_data else 0,
                'name': service_detail_val.name,
                'description': service_detail_val.description
            },
            'locations':
                list(service_user_val),
            'path': path_data
        }
    except User_Enterprise.DoesNotExist:
        pass
    except Exception as error:
        response['message'] = 'Algo salio mal, contacta al administrador para solucionarlo ' + str(error)

    return Response(response, status_response)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_location_trace(request, pk):
    response = {}
    status_response = status.HTTP_400_BAD_REQUEST
    response['status'] = False
    try:
        user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
        service_user_val = Service_User.objects.filter(service_id=pk, state=True).values_list('id', flat=True)

        try:
            user_val = Service_User.objects.get(service_id=pk, state=True).user
            user_data = {
                'name': user_val.first_name + ' ' + user_val.first_last_name,
                'identification': user_val.type_identification.name + ' ' + str(user_val.identification),
                'phone': user_val.phone,
            }
        except:
            user_data = {}

        service_user_ids = list(service_user_val)
        service_location_id = pk

        trazability_vals = Service_Trazability.objects.filter(Q(group=2, reference=service_location_id) | Q(group=3, reference__in=service_user_ids)).order_by('creation_date').select_related('process_state').values('process_state__name', 'date_trace', 'hour_trace')

        response['status'] = True
        response['data'] = {
            'user': user_data,
            'trace': list(trazability_vals),
        }
        status_response = status.HTTP_200_OK
    except User_Enterprise.DoesNotExist:
        pass
    except Exception as error:
        response['message'] = 'Algo salio mal, contacta al administrador para solucionarlo ' + str(error)

    return Response(response, status_response)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def route_services_offline(request):
    response = {}
    response['status'] = False
    route_list = []
    try:
        data = request.data
        route_service_val = Service_Location.objects.filter(service_id = data['service_id']
            ).select_related('project').values(
                'id',
                'comment',
                'comment_address',
                'user_name',
                'user_email',
                'address_normalized',
                'guide_number',
                'user_phone_number',
                'order',
                'latitude',
                'longitude',
                'state',
                'process_state_id',
                'state',
                'service_id',
                'project_id',
                'project__form_id',
            )
        # print(len(route_service_val))
        if len(route_service_val) != 0:
            for route in route_service_val:
                route_list.append({
                    'id': route['id'],
                    'comment': route['comment'],
                    'comment_address': route['comment_address'],
                    'user_name' : route['user_name'],
                    'user_email' : route['user_email'],
                    'address_normalized' : route['address_normalized'],
                    'guide_number' : route['guide_number'],
                    'user_phone_number' : route['user_phone_number'],
                    'order' : route['order'],
                    'latitude' : route['latitude'],
                    'longitude' : route['longitude'],
                    'state' : route['state'],
                    'process_state_id' : route['process_state_id'],
                    'service_id' : route['service_id'],
                    'project_id' : route['project_id'],
                    'project__form' : route['project__form_id'],
                })
            response['status'] = True
            response['route'] = route_list
        else:
            response['status'] = False
            response['message'] = "No hay rutas para mostrar"
    except Service_Location.DoesNotExist:
        response['message'] = 'No hay rutas para mostrar'
    except Exception as error:
        print(error)
        response['message'] = 'Algo salio mal, verifica la información enviada ' + str(error)
    return Response(response)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_file_from_route(request):
    response = {}
    status_response = status.HTTP_400_BAD_REQUEST
    response['status'] = False
    data = request.data
    try:
        user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
        route_val_id = int(data['id'])
        route_val = Route.objects.get(massive_file_id=route_val_id)
        file_type = int(data['type'])
        file_data = None
        if file_type in [0,1]:
            column_vals = [
                'Direccion',
                'Nombre Destinatario',
                'Documento Destinatario',
                'Numero Celular',
                'Email',
                'Ciudad',
                'Departamento',
                'Peso del Producto',
                'Numero de Guia',
                'Observaciones',
                'Orden',
                'Grupo',
                'Proyecto'
            ]

            name_file = ''

            if file_type == 0:
                file_path = os.path.join(settings.MEDIA_ROOT, str(user_val.enterprise_id), str(route_val_id), 'Subidas.xlsx')

                try:
                    content = None
                    with open(file_path, 'rb') as f:
                        content = f.read()
                    response = HttpResponse(content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename="{}"'.format('Subidas.xlsx')
                    return response
                except FileNotFoundError:
                    pass
                # except Exception as error:
                #     print(type(error), error)
                #     return response

                file_path = '{}{}'.format(settings.MEDIA_ROOT, str(route_val.json_path))
                json_data = []
                with open(file_path, 'r') as f:
                    json_data = json.loads(f.read())

                file_data = [
                    [
                        value['address'],
                        value['addressee'],
                        value['id_number'],
                        value['phone'],
                        value['email'],
                        value['city'],
                        value['department'],
                        value['weight'],
                        value['number_guide'],
                        value['observation'],
                        value['order'],
                        value['group'],
                        value['project'],
                        value['lat'],
                        value['lon'],
                    ]
                    for value in json_data
                ]

                if len(file_data):
                    column_vals.append('Latitud')
                    column_vals.append('Longitud')

                    name_file = 'Subidas.xlsx'

            elif file_type == 1:
                file_path = os.path.join(settings.MEDIA_ROOT, str(user_val.enterprise_id), str(route_val_id), 'Error.xlsx')

                try:
                    content = None
                    with open(file_path, 'rb') as f:
                        content = f.read()
                    response = HttpResponse(content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename="{}"'.format('Error.xlsx')
                    return response
                except FileNotFoundError:
                    pass
                except Exception as error:
                    print(type(error), error)
                    return response

                massive_file_val = route_val.massive_file

                massive_errors_vals = Massive_Errors.objects.filter(massive_file_id=massive_file_val.id).order_by('row').values('row', 'data')
                massive_errors_vals = {
                    error['row']:error['data'].split(' - ')[0]
                    for error in massive_errors_vals
                }
                error_rows = massive_errors_vals.keys()

                file_path = '{}/{}'.format(settings.MEDIA_ROOT, str(massive_file_val.template))
                data_frame = pd.read_excel(file_path)
                row_values = data_frame.to_numpy().tolist()
                file_data = []
                for index, row in enumerate(row_values):
                    temp = row
                    key = index + 1
                    if key in error_rows:
                        temp.append(massive_errors_vals[key])
                        file_data.append(temp)
                    pass
                if len(file_data):
                    column_vals.append('Error')
                    name_file = 'Error.xlsx'

            if name_file:
                file_path = os.path.join(settings.MEDIA_ROOT, str(user_val.enterprise_id), str(route_val_id))

                if not os.path.exists(file_path):
                    os.makedirs(file_path)

                file_path = os.path.join(file_path, name_file)

                data_frame = pd.DataFrame(file_data, columns=column_vals)

                data_frame.to_excel(file_path, index=False)

                try:
                    content = None
                    with open(file_path, 'rb') as f:
                        content = f.read()
                    response = HttpResponse(content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename="{}"'.format('Error.xlsx')
                    return response
                except FileNotFoundError:
                    pass
                except Exception as error:
                    print(type(error), error)
                    return response

        elif file_type == 2:
            massive_file_val = route_val.massive_file
            file_path = '{}/{}'.format(settings.MEDIA_ROOT, str(massive_file_val.template))
            file_name = os.path.basename(file_path)
            file_bytes = open(file_path, 'rb')
            response = HttpResponse(file_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="{}"'.format(file_path)
            return response

    except User_Enterprise.DoesNotExist:
        pass

    return Response(response)
