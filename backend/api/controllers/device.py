from api.models import User_Enterprise, Type_Device, Device_Enterprise, Device_Registry, Massive_Device, Massive_File, Massive_Errors
from rest_framework.views import APIView
from api.serializers import DeviceSerializer, TypeDeviceSerializer, MassiveFileSerializer
from api.permissions import IsSuperAdmin, IsAdmin, IsUserAdminOrHasPermission
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import Http404, HttpResponseBadRequest
from api.controllers.admin import get_enterprise
from datetime import datetime
from threading import Thread
import openpyxl
import requests

class TypeDeviceList(APIView):
    """
    API endpoint que permite la consulta de los permisos y la creación de nuevos.
    """
    permission_classes = [IsUserAdminOrHasPermission]

    # Consulta
    def get(self, request, format=None):
        try:
            enterprise = get_enterprise(request)
            type_devices = Type_Device.objects.filter(enterprise_id=enterprise)
            serializer = TypeDeviceSerializer(type_devices, many=True)
            return Response({'status': True, 'data': serializer.data})
        except User_Enterprise.DoesNotExist:
            return HttpResponseBadRequest({'status': False, 'detail': "This user isn't on the system."})
        
    def post(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            data = request.data
            data['enterprise_id'] = enterprise
            
            try:
                Type_Device.objects.get(name=data['name'], enterprise_id=enterprise)
                response['message'] = 'Ya hay una categoria dispositivo con este nombre'
            except Type_Device.DoesNotExist:    
                type_device_new = Type_Device()
                type_device_new.name = data['name']
                type_device_new.description = data['description']
                type_device_new.enterprise_id = data['enterprise_id']
                type_device_new.icon = data['icon']
                type_device_new.save()
                
                data['id'] = type_device_new.id
                del data['enterprise_id']
                response['status'] = True
                response['data'] = data
                status_response = status.HTTP_201_CREATED
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status=status_response)


class TypeDeviceDetail(APIView):
    permission_classes = [IsUserAdminOrHasPermission]
    
    def get_object(self, pk, enterprise_id):
        try:
            return Type_Device.objects.get(pk=pk, enterprise_id=enterprise_id)
        except Type_Device.DoesNotExist:
            raise Http404

    def get(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            type_device = self.get_object(pk, enterprise)
            serializer = TypeDeviceSerializer(type_device)
            response['status'] = True
            response['data'] = serializer.data
            status_response = status.HTTP_200_OK
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status=status_response)

    def put(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        
        try:
            enterprise = get_enterprise(request)
            data = request.data
            try:
                Type_Device.objects.exclude(id=pk).get(name=data['name'], enterprise_id=enterprise)
                response['message'] = 'Ya hay una categoria dispositivo con este nombre'
            except Type_Device.DoesNotExist:
                type_device = self.get_object(pk, enterprise)
                type_device.name = data['name']
                type_device.description = data['description']
                type_device.icon = data['icon']
                type_device.state = data['state']
                type_device.save()
                
                response['status'] = True
                response['data'] = data
                status_response = status.HTTP_202_ACCEPTED
            except Type_Device.MultipleObjectsReturned:
                response['message'] = 'Ya hay una categoria dispositivo con este nombre'
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status=status_response)

    def delete(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            type_device = self.get_object(pk, enterprise)
            if Device_Enterprise.objects.filter(enterprise_id=enterprise, type_device=type_device).count() > 0:
                type_device.state = False
                type_device.save()
                status_response=status.HTTP_202_ACCEPTED
                response['status'] = True
                response['message'] = 'Esta categoria ha sido inhabilitada, debido a que existen dispositivos vinculados a esta.'
            else:    
                type_device.delete()
                status_response=status.HTTP_204_NO_CONTENT
                response['status'] = True
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status=status_response)


class DeviceList(APIView):
    """
    API endpoint que permite la consulta de los permisos y la creación de nuevos.
    """
    permission_classes = [IsUserAdminOrHasPermission]

    # Consulta
    def get(self, request, format=None):
        try:
            enterprise = get_enterprise(request)
            devices = Device_Enterprise.objects.filter(enterprise_id=enterprise, state=True).order_by('creation_date')
            serializer = DeviceSerializer(devices, many=True)
            return Response({'status': True, 'data': serializer.data})
        except User_Enterprise.DoesNotExist:
            return HttpResponseBadRequest({'status': False, 'detail': "This user isn't on the system."})
    
    def post(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            data = request.data
            data['enterprise_id'] = enterprise
            
            device_new = Device_Enterprise()
            device_new.name = data['name']
            device_new.type_device_id = data['type_device_id']
            device_new.enterprise_id = data['enterprise_id']
            device_new.mac = data['mac']
            device_new.state = data['state'] == 'true'
            device_new.save()
            
            data['id'] = device_new.id
            del data['enterprise_id']
            response['status'] = True
            response['data'] = data
            status_response = status.HTTP_201_CREATED
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status=status_response)

class DeviceDetail(APIView):
    permission_classes = [IsUserAdminOrHasPermission]
    
    def get_object(self, pk, enterprise_id):
        try:
            return Device_Enterprise.objects.get(pk=pk, enterprise_id=enterprise_id)
        except Type_Device.DoesNotExist:
            raise Http404

    def get(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            device = self.get_object(pk, enterprise)
            serializer = DeviceSerializer(device)
            response['status'] = True
            response['data'] = serializer.data
            status_response = status.HTTP_200_OK
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status=status_response)

    def put(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        
        try:
            enterprise = get_enterprise(request)
            data = request.data
            
            device = self.get_object(pk, enterprise)
            device.name = data['name']
            device.type_device_id = data['type_device_id']
            device.mac = data['mac']
            device.modify_date = datetime.now()
            device.state = data['state']
            device.save()
            
            response['status'] = True
            response['data'] = data
            status_response = status.HTTP_202_ACCEPTED
            
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status=status_response)

    def delete(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            device = self.get_object(pk, enterprise)
            device.state = False
            device.save()
            
            status_response=status.HTTP_204_NO_CONTENT
            response['status'] = True
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status=status_response)

class MassiveDevice(APIView):

    permission_classes = [IsUserAdminOrHasPermission]
    
    def get(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            mass_files = Massive_File.objects.filter(enterprise_id=enterprise, type=2).order_by('date')
            serializer = MassiveFileSerializer(mass_files, many=True)
            return Response({'status': True, 'data': serializer.data})
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)

    def post(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST

        if "template" in request.FILES:
            try:
                enterprise = get_enterprise(request)
                name_file = str(request.FILES['template'])
                if name_file.lower().endswith(('.xlsx')):
                    xlsx_file = request.FILES['template']
                    wb_obj = openpyxl.load_workbook(xlsx_file)

                    sheet = wb_obj.active

                    print('sheet.max_row, sheet.max_column')
                    print(sheet.max_row, sheet.max_column)

                    if sheet.max_row > 1:
                        amount_rows = sheet.max_row - 1
                        massive_val = Massive_File()
                        massive_val.enterprise_id = enterprise
                        massive_val.amount = amount_rows
                        massive_val.type = 2
                        massive_val.template = request.FILES['template']
                        massive_val.save()
                        new_thread = Thread(target=run_device, args=(massive_val, sheet,))
                        new_thread.start()
                        response['status'] = True
                        response['data'] = {'id': massive_val.id, 'template': name_file, 'amount': amount_rows, 'success': 0, 'errors':0, 'progressbar':0, 'status':0}
                        status_response = status.HTTP_202_ACCEPTED
            except User_Enterprise.DoesNotExist:
                pass
        return Response(response, status_response)
        
def run_device(massive, sheet):
    
    categories_values = Type_Device.objects.filter(enterprise_id=massive.enterprise_id)
    categories = {category.name : category.id for category in list(categories_values)}
    
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=(massive.amount + 1), max_col=5)):
        num_row = i + 2
        name = ''
        category = ''
        mac = ''
        for cell in row:
            if cell.value is not None:
                if cell.column == 1:
                    name = cell.value.strip()
                elif cell.column == 2:
                    category = cell.value.strip()
                elif cell.column == 3:
                    mac = cell.value.strip().upper()
        try:
            error = ''
            if name == '' or mac == '':
                raise Exception('Datos incompletos')
            else:
                tem_val = name.replace(' ', '')
                if not tem_val:
                    error += (", " if error != "" else "" ) + 'Nombre'
                tem_val = category.replace(' ', '')
                if not tem_val.replace(' ', ''):
                    error += (", " if error != "" else "" ) + 'Tipo Dispositivo'
                tem_val = mac.replace(' ', '')
                if not tem_val.replace(' ', ''):
                    error += (", " if error != "" else "" ) + 'MAC'
                if error != '':
                    error += (' invalidos' if ',' in error else ' invalido')
                    raise Exception(error)

            device_val = Device_Enterprise()
            
            device_val.mac = mac
            device_val.name = name
            device_val.enterprise_id = massive.enterprise_id
            if category in categories:
                device_val.type_device_id = categories[category]
            else:
                raise Exception('Dato tipo dispositivo no coincide con ninguna de las opciones ' + (", ".join(categories.keys())))
            
            device_val.save()

            massive_val = Massive_Device()
            massive_val.massive_file = massive
            massive_val.device = device_val
            massive_val.save()

            massive.success += 1
            massive.save()
        except Exception as e:
            massive.error += 1
            data_error = (str(e) +
                ' - nombre: ' + name +
                ', mac: ' + mac +
                ', categoria: ' + category
            )
            # print(data_error)
            error_val = Massive_Errors()
            error_val.massive_file = massive
            error_val.row = num_row
            error_val.data = data_error
            error_val.save()
    massive.status = 1
    massive.save()

class MassiveDeviceErrors(APIView):
    
    permission_classes = [IsUserAdminOrHasPermission]
    
    def get(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        errors = []
        try:
            enterprise = get_enterprise(request)
            massive_errors_values = Massive_Errors.objects.filter(massive_file_id=pk, massive_file__type=2, massive_file__enterprise_id=enterprise)
            errors = []
            for massive_errors_val in massive_errors_values:
                error = {
                    'row': massive_errors_val.row,
                    'data': massive_errors_val.data
                }
                errors.append(error)
            response['status'] = True
            response['data'] = errors
            status_response = status.HTTP_200_OK
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)
    
class DeviceData(APIView):
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        enterprise = get_enterprise(request)
        print(enterprise)
        if enterprise in [1,2,3,5]:
            data = '''
            union(
                tables:[
                    from(bucket: "Devices")
                        |> range(start: -30d)
                        |> filter(fn: (r) => r["_measurement"] == "cpuinfo")
                        |> filter(fn: (r) => r["_field"] == "siblings")        
                        |> sort(columns: ["_time"], desc: true)
                        |> unique(column: "serial_no")        
                        |> group(columns: ["host", "_measurement", "serial_no"], mode: "by")        
                        |> max()        
                        |> yield(name: "max_core")        
                        |> rename(columns: {"_value":"core_max"}),     
                    from(bucket: "Devices")        
                        |> range(start: -30d)        
                        |> filter(fn: (r) => r["_measurement"] == "lscpu")        
                        |> filter(fn: (r) => r["_field"] == "model_name")        
                        |> sort(columns: ["_time"], desc: true)
                        |> unique(column: "serial_no")        
                        |> group(columns: ["host", "_measurement", "serial_no"], mode: "by")        
                        |> yield(name: "cpu_model"),      
                    from(bucket: "Devices")    
                        |> range(start: -30d)    
                        |> filter(fn: (r) => r["_measurement"] == "proc_meminfo")        
                        |> filter(fn: (r) => r["_field"] == "MemTotal")        
                        |> sort(columns: ["_time"], desc: true)
                        |> unique(column: "serial_no")        
                        |> group(columns: ["serial_no"])        
                        |> rename(columns: {_value: "Ram"})        
                        |> yield(name: "ram"),      
                    from(bucket: "Devices")        
                        |> range(start: -30d)        
                        |> filter(fn: (r) => r["_measurement"] == "disks")        
                        |> filter(fn: (r) => r["_value"] != -1)        
                        |> sort(columns: ["_time"], desc: true)
                        |> unique(column: "serial_no")         
                        |> yield(name: "disk"),      
                    from(bucket: "Devices")        
                        |> range(start: -30d)        
                        |> filter(fn: (r) => r["_measurement"] == "geo")        
                        |> filter(fn: (r) => r["_field"] == "latitud" or r["_field"] == "longitud")        
                        |> filter(fn: (r) => r["_value"] != "NaN")        
                        |> sort(columns: ["_time"], desc: true)
                        |> unique(column: "serial_no")        
                        |> yield(name: "geo")    
                ]
            )'''
            r = requests.post("http://habilidapp.com:8086/api/v2/query?orgID=fc866c257b605b50", data, headers={
                'Accept': 'text/csv',
                'Authorization': 'Token bwZyVkZSypvTxAbGb-MRP8e_7GHKjoqVFGZbk9EJWcXcxs--I5NERTMkuDvyiMWQaxLVJ9aU0Yd8heoLQ6xrNw==',
                'Content-Type': 'application/vnd.flux'
                },verify=False)
            return Response({'data':r.text})
        else:
            return Response({'data':''})