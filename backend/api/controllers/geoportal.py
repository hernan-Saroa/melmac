from ast import Or
from api.controllers.admin import get_enterprise
from api.controllers.api import check_usage_api_enterprise, register_api_usage
from api.controllers.routing import geocode_address
from api.models import Massive_Errors, Massive_File, User_Enterprise, Geo_Portal, Follow_User, Follow_User_Offline, Traceability_User
from api.controllers.admin import get_enterprise
from api.permissions import IsUserAdminOrHasPermission
from api.util import AddressNormalizer
from django.conf import settings
from django.db.models import Q
from geopy.geocoders import GoogleV3
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from threading import Thread
from datetime import datetime, timedelta

import itertools
import json
import openpyxl
import pytz
import os

NORMALIZER_INSTANCE = AddressNormalizer()


TZ_INFO = pytz.timezone('America/Bogota')


# Lista de Hilos para verificar el estado de los mismos.
threads_list = {}

# models
from django.db.models import F, Value
from api.models import (
    Answer_Consecutive,
    Answer_Form,
    Geo_Portal,
    User_Enterprise,
)
# others
from itertools import chain
import operator
import json

class GeoPortalPointList(APIView):
    # permission_classes = [IsAuthenticated]
    permission_classes = [IsUserAdminOrHasPermission]

    def get(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)

            geoportal_vals = Geo_Portal.objects.filter(
                massive_file__enterprise_id=enterprise,
                state=True
                ).select_related(
                    'massive_file',
                    'created_by'
                ).order_by(
                    '-massive_file__date'
                ).values(
                    'id',
                    'massive_file__amount',
                    'massive_file__success',
                    'massive_file__error',
                    'massive_file__date',
                    'name',
                    'description',
                    'json_path',
                    'created_by__first_name',
                    'created_by__middle_name',
                    'created_by__first_last_name',
                    'created_by__second_last_name',
                    'created_by__email',
                    'status',
                )

            geoportal_values = []
            for geoportal in geoportal_vals:
                try:
                    path = settings.MEDIA_ROOT + str(geoportal['json_path'])
                    content = ''
                    with open(path, 'r') as f:
                        content = f.read()
                    data_json = json.loads(content)
                    addresses = []
                    if data_json and 'addresses' in data_json:
                        # print(data_json['addresses'].keys())
                        for key in list(data_json['addresses'].keys()):
                            for addr in data_json['addresses'][key]:
                                addresses.append(addr)
                        geoportal['json_path'] = addresses

                        if 'list' not in request.data:
                            geoportal_values.append(geoportal)

                except FileNotFoundError:
                    pass

                if 'list' in request.data:
                    geoportal_values.append(geoportal)
            response = {
                'status': True,
                'data': geoportal_values
            }
            status_response = status.HTTP_200_OK
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)

    def post(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST

        data = request.POST
        # print(data)
        try:
            user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
            enterprise = user_val.enterprise_id
            if not 'name' in data or not 'description' in data or not 'template' in request.FILES:
                response['detail'] = 'Hacen falta datos para registrar este proceso'
                return Response(response, status_response)

            if (data['name'] and
                data['name'].strip() != '' and
                data['description'] and
                data['description'].strip() != ''):

                geoportal_val = Geo_Portal()

                geoportal_val.name = data['name']
                geoportal_val.description = data['description']
                geoportal_val.created_by = user_val


                if "template" in request.FILES:
                    name_file = str(request.FILES['template'])
                    if name_file.lower().endswith('.xlsx'):
                        xlsx_file = request.FILES['template']
                        wb_obj = openpyxl.load_workbook(xlsx_file)

                        sheet = wb_obj.active

                        massive_val = Massive_File()
                        massive_val.enterprise_id = enterprise
                        massive_val.amount = sheet.max_row - 1
                        massive_val.type = 5
                        massive_val.template = request.FILES['template']
                        massive_val.save()

                        geoportal_val.massive_file = massive_val
                        geoportal_val.save()

                        if sheet.max_row > 1:

                            thread_ID = 'TH-'+str(enterprise)+'-'+str(geoportal_val.id)

                            new_thread = PointsLoadingThread(name_file, thread_ID, sheet, geoportal_val)
                            new_thread.start()

                            response['status'] = True
                            response['data'] = {
                                'id': geoportal_val.id,
                                'name': geoportal_val.name,
                                'description': geoportal_val.description,
                                'massive_file__amount': massive_val.amount,
                                'massive_file__success': 0,
                                'massive_file__error': 0,
                                'massive_file__date': massive_val.date,
                                'created_by': '{} {}'.format(user_val.first_name, user_val.first_last_name),
                                'status': 0
                            }
                            status_response = status.HTTP_202_ACCEPTED
        except User_Enterprise.DoesNotExist:
            pass
        except TypeError:
            response['detail'] = 'Este archivo presenta problemas para ser procesado, por favor vuelvalo a generar.'
        except Exception as e:
            response['detail'] = str(e)
        return Response(response, status_response)

class PointsLoadingThread(Thread):

    def __init__(self, thread_name, thread_ID, sheet, geoportal_val):
        Thread.__init__(self)
        self.thread_name = thread_name
        self.thread_ID = thread_ID
        self.sheet = sheet
        self.daemon = True
        self.done = False
        self.json_data = []
        self.geoportal = geoportal_val

        # helper function to execute the threads
    def run(self):
        print(str(self.thread_name) +" "+ str(self.thread_ID))
        self.run_route(self.sheet)

    def run_route(self, sheet):
        massive = self.geoportal.massive_file
        enterprise = massive.enterprise_id
        self.total = sheet.max_row - 1
        error_count = 0
        loc_file_name = '/dir.json'
        path = settings.MEDIA_ROOT

        # try:
        if not os.path.exists(path):
            os.makedirs(path)
        path += loc_file_name

        try:
            content_file = ""
            with open(path, "r") as f:
                content_file = f.read()
            loaded_addresses = json.loads(content_file)
        except:
            loaded_addresses = []
        list_addresses = { addr['address'] + ',' + addr['city']: addr for addr in loaded_addresses if addr['address'] and addr['city']}
        list_coordinates = { '{},{}'.format(addr['lat'], addr['lon']): addr for addr in loaded_addresses }
        data = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=6)):
            num_row = i + 2

            # columns data
            name = ''
            address = ''
            city = ''
            department = ''
            latitude = ''
            longitude = ''
            try:
                for cell in row:
                    if cell.value is not None:
                        if cell.column == 1:
                            name = cell.value.strip()
                        elif cell.column == 2:
                            address = cell.value.strip()
                        elif cell.column == 3:
                            city = cell.value.strip()
                        elif cell.column == 4:
                            department = cell.value.strip()
                        elif cell.column == 5:
                            # latitude = float(str(cell.value).strip()) if cell.value else ''
                            latitude = str(cell.value).strip()
                        elif cell.column == 6:
                            # longitude = float(str(cell.value).strip()) if cell.value else ''
                            longitude = str(cell.value).strip()
            except Exception as e:
                print("error12345")
                print(e)
            try:
                add = False
                if address == '' or city == '' or name == '':
                    raise Exception('error con los campos: {}'.format(address, city, name))

                # Proceso para direcciones
                if latitude == '' and longitude == '':
                    address_normalize = AddressNormalizer().normalize(address)
                    if address_normalize:
                        if address_normalize[0]['address'] == "" and address_normalize[1]['address'] == "":
                            raise Exception('La dirección no pudo normalizarse')

                    search_address = address_normalize[0]['address'] if address_normalize[0]['address'] != "" else address_normalize[1]['address']
                    key = '{},{}'.format(address, city)
                    if key not in list_addresses:
                        if check_usage_api_enterprise(1, enterprise):

                            address_info = geocode_address(search_address, city)

                            register_api_usage(enterprise, 1, 500 if not address_info else 200)

                            if not address_info:
                                raise Exception('La dirección no se encontro')

                            temp = {
                                'address': address,
                                'address_normalize': address_normalize,
                                'city': city,
                                'department': address_info['administrative_area_level_1'],
                                'lat': address_info['lat'],
                                'lon': address_info['lng'],
                                'name': name,
                                'zone': address_info,
                            }

                            add = True
                        else:
                            raise Exception('Ya no tienes mas usos de la API, obten un nuevo paquete e intenta de nuevo.')
                    else:
                        temp = list_addresses[key]
                        temp['name'] = name
                else:
                    # Proceso para coordenadas
                    if latitude == '' or longitude == '':
                        raise Exception('Datos incompletos - Lat,Lng')
                    
                    isLatitude = is_float(latitude)
                    isLongitude = is_float(longitude)
                    if isLatitude == False or isLongitude == False:
                        raise Exception('Latitud o Longitud no válida lat: {}, lon: {}'.format(latitude, longitude))
                    
                    latitude = float(latitude)
                    longitude = float(longitude)
                    key = '{},{}'.format(latitude, longitude)
                    if key not in list_coordinates:
                        address_info = None
                        # address_info = reverse_geocode([latitude, longitude])
                        # if not address_info:
                        #     raise Exception('La dirección de las coordenadas no se encontro')

                        temp = {
                            'address': address,
                            'address_normalize': address_normalize if address_info else None,
                            'city': city,
                            'department': department,
                            'lat': latitude,
                            'lon': longitude,
                            'name': name,
                            'zone': address_info,
                        }

                        add = True
                    else:
                        temp = list_coordinates[key]

                data.append(temp)

                if add:
                    loaded_addresses.append(temp)
                massive.success += 1
            except Exception as e:
                # response['message'] = 'Hay error en algunos campos del excel: ' + str(e)
                # return Response(response)
                print ("error cargando111")
                print(e)
                massive.error += 1
                data_error = (str(e),
                    'Dirección: ' + address +
                    ', ciudad: ' + city +
                    ', departamento: ' + department
                )
                error_val = Massive_Errors()
                error_val.massive_file = massive
                error_val.row = num_row
                error_val.data = data_error
                error_val.save()
                error_count += 1
            massive.save()

        self.json_data = {
            'loaded': len(data),
            'errors': error_count
        }

        get_attr = lambda point: point['city'] if 'city' in point and point['city'] else 'N/A'
        points_list = {k: list(g) for k, g in itertools.groupby(sorted(data, key=get_attr), get_attr)}
        self.json_data['addresses'] = points_list

        with open(path, "w") as f:
            f.write(json.dumps(loaded_addresses))

        self.done = True
        self.geoportal.status = 1
        # except:
        #     self.geoportal.status = 3

        loc_file_name = 'data_' + str(self.geoportal.id) + '.json'
        folder = '/' + str(massive.enterprise_id) + '/geoportal/'
        path = settings.MEDIA_ROOT + folder

        if not os.path.exists(path):
            os.makedirs(path)
        path += loc_file_name

        with open(path, "w+") as destination:
            destination.write(json.dumps(self.json_data))

        self.geoportal.json_path = folder + loc_file_name

        self.geoportal.save()

def is_float(string):
    if '.' in string: 
        try:
            float(string)
            return True
        except ValueError:
            return False
    else:
        return False

def reverse_geocode(coordinates):
    # geolocator = Nominatim(user_agent="melmac")
    # location = geolocator.geocode(address + ", " + city + ", colombia", addressdetails=True, language="es")

    geolocator = GoogleV3(api_key="AIzaSyCROv_TbEcFuZ-BOX4kGH4-PBinysfc6m0")
    location = geolocator.reverse(coordinates, language="es")

    if location:

        # print(coordinates, location, location.raw)
        response = {}
        response['street_address'] = location.raw['formatted_address']
        response['neighborhood'] = ''
        response['locality'] = ''
        response['sublocality'] = ''
        response['postal_code'] = ''
        response['administrative_area_level_1'] = ''
        response['country'] = ''
        response['lat'] = location.raw['geometry']['location']['lat']
        response['lng'] = location.raw['geometry']['location']['lng']

        keys = ['neighborhood', 'locality', 'sublocality', 'postal_code', 'administrative_area_level_1', 'country']
        for key in keys:
            for part in location.raw['address_components']:
                if key in part['types']:
                    response[key] = part['long_name']

        response['address'] = response['street_address']
        response['locality'] = NORMALIZER_INSTANCE.remove_accent_mark(response['locality']).capitalize()

        return response

    return location


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def check_status_points(request):
    response = {
        'status': False
    }
    status_response = status.HTTP_400_BAD_REQUEST

    try:

        if 'id' in request.data:
            geoportal_id = request.data['id']
            total = 0
            processed = 0

            geoportal_val = Geo_Portal.objects.get(id=geoportal_id, state=True)

            geo_status = geoportal_val.status

            total = geoportal_val.massive_file.amount
            processed = geoportal_val.massive_file.success + geoportal_val.massive_file.error

            if geoportal_val.json_path:

                path = settings.MEDIA_ROOT + geoportal_val.json_path

                with open(path, 'r') as f:
                    json_data = json.loads(f.read())

                response['data'] = json_data

            response['status'] = processed == total
            response['total'] = total
            response['processed'] = processed
            response['state'] = geo_status

            status_response = status.HTTP_200_OK
        else:
            response['detail'] = 'Parametros incorrectos.'

    except Geo_Portal.DoesNotExist as e:
        response['detail'] = 'No se encuentra el proceso.'
    except Exception as e:
        print(e)
        response['detail'] = 'Un error inesperado ha sucedido.'
        status_response = status.HTTP_500_INTERNAL_SERVER_ERROR

    return Response(response, status_response)

class GeoPortalAnswerList(APIView):
    permission_classes = [IsAuthenticated]
    # permission_classes = [IsUserAdminOrHasPermission]

    def get(self, request, position, format=None):
        return self.query(request, 'GET', position)

    def post(self, request, position, format=None):
        return self.query(request, 'POST', position)

    def query(self, request, method, position):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        filter_data = None
        filter = False

        if method == 'POST':
            filter = True
            filter_data = request.data
            if filter_data['doc'] and len(filter_data['doc']) == 0:
                response = {"status": True, 'data': []}
                status_response = status.HTTP_200_OK

                return Response(response, status_response)

        try:
            user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)

            date_ini = None
            date_fin = None
            doc_values = None
            usr_ids = None
            public = False
            role_val = None

            if filter:
                if filter_data['date_ini']:
                    date_ini = datetime.strptime(filter_data['date_ini'], "%Y-%m-%d")
                    date_ini = date_ini.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=TZ_INFO)
                if filter_data['date_fin']:
                    date_fin = datetime.strptime(filter_data['date_fin'], "%Y-%m-%d") + timedelta(days=1)
                    date_fin = date_fin.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=TZ_INFO)
                if filter_data['doc']:
                    doc_values = [int(doc) for doc in filter_data['doc']]
                if filter_data['user']:
                    usr_ids = filter_data['user']
                    if 'null' in usr_ids:
                        idx = usr_ids.index('null')
                        del usr_ids[idx]
                        public = True
                elif filter_data['role']:
                    role_val = filter_data['role']
            # Los 3 Filtros
            if date_ini and date_fin and doc_values:
                answer_base = Answer_Form.objects.filter(
                    form_enterprise__enterprise_id=user_val.enterprise_id,
                    form_enterprise__state=True,
                    consecutive=False,
                    creation_date__gte=date_ini,
                    creation_date__lt=date_fin,
                    form_enterprise_id__in=doc_values,
                    state=True
                )
            # Fecha Inicial y Documento(s)
            elif date_ini and doc_values:
                date_fin = (datetime.now(tz=TZ_INFO)).replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)

                answer_base = Answer_Form.objects.filter(
                    form_enterprise__enterprise_id=user_val.enterprise_id,
                    form_enterprise__state=True,
                    consecutive=False,
                    creation_date__gte=date_ini,
                    creation_date__lt=date_fin,
                    form_enterprise_id__in=doc_values,
                    state=True
                )
            # Fecha Inicial y Final
            elif date_ini and date_fin:
                answer_base = Answer_Form.objects.filter(
                    form_enterprise__enterprise_id=user_val.enterprise_id,
                    form_enterprise__state=True,
                    consecutive=False,
                    creation_date__gte=date_ini,
                    creation_date__lt=date_fin,
                    state=True
                )
            # Fecha Final y Documento(s)
            elif date_fin and doc_values:
                answer_base = Answer_Form.objects.filter(
                    form_enterprise__enterprise_id=user_val.enterprise_id,
                    form_enterprise__state=True,
                    consecutive=False,
                    creation_date__lt=date_fin,
                    form_enterprise_id__in=doc_values,
                    state=True
                )
            # Dia especifico
            elif date_ini:
                date_fin = (datetime.now(tz=TZ_INFO)).replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
                answer_base = Answer_Form.objects.filter(
                    form_enterprise__enterprise_id=user_val.enterprise_id,
                    form_enterprise__state=True,
                    consecutive=False,
                    creation_date__gte=date_ini,
                    creation_date__lt=date_fin,
                    state=True
                )
            elif date_fin:
                answer_base = Answer_Form.objects.filter(
                    form_enterprise__enterprise_id=user_val.enterprise_id,
                    form_enterprise__state=True,
                    consecutive=False,
                    creation_date__lt=date_fin,
                    state=True
                )
            # Documento(s)
            elif doc_values:
                answer_base = Answer_Form.objects.filter(
                    form_enterprise__enterprise_id=user_val.enterprise_id,
                    form_enterprise__state=True,
                    consecutive=False,
                    form_enterprise_id__in=doc_values,
                    state=True
            )
            else:
                answer_base = Answer_Form.objects.filter(
                    form_enterprise__enterprise_id=user_val.enterprise_id,
                    form_enterprise__state=True,
                    consecutive=False,
                    state=True
                )

            # 3 Filtros
            if date_ini and date_fin and doc_values:
                consecutive_base = Answer_Consecutive.objects.filter(
                    form_consecutive__enterprise_id=user_val.enterprise_id,
                    form_consecutive__state=True,
                    creation_date__gte=date_ini,
                    creation_date__lt=date_fin,
                    form_consecutive_id__in=doc_values,
                    state=True
                )
            # Fecha Inicial y Documento(s)
            elif date_ini and doc_values:
                date_fin = (datetime.now(tz=TZ_INFO)).replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)

                consecutive_base = Answer_Consecutive.objects.filter(
                    form_consecutive__enterprise_id=user_val.enterprise_id,
                    form_consecutive__state=True,
                    creation_date__gte=date_ini,
                    creation_date__lt=date_fin,
                    form_consecutive_id__in=doc_values,
                    state=True
                )
            # Fecha Inicial y Final
            elif date_ini and date_fin:
                consecutive_base = Answer_Consecutive.objects.filter(
                    form_consecutive__enterprise_id=user_val.enterprise_id,
                    form_consecutive__state=True,
                    creation_date__gte=date_ini,
                    creation_date__lt=date_fin,
                    state=True
                )
            # Fecha Final y Documento(s)
            elif date_fin and doc_values:
                consecutive_base = Answer_Consecutive.objects.filter(
                    form_consecutive__enterprise_id=user_val.enterprise_id,
                    form_consecutive__state=True,
                    creation_date__lt=date_fin,
                    form_consecutive_id__in=doc_values,
                    state=True
                )
            # Dia especifico
            elif date_ini:
                date_fin = (datetime.now(tz=TZ_INFO)).replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
                consecutive_base = Answer_Consecutive.objects.filter(
                    form_consecutive__enterprise_id=user_val.enterprise_id,
                    form_consecutive__state=True,
                    creation_date__gte=date_ini,
                    creation_date__lt=date_fin,
                    state=True
                )
            elif date_fin:
                consecutive_base = Answer_Consecutive.objects.filter(
                    form_consecutive__enterprise_id=user_val.enterprise_id,
                    form_consecutive__state=True,
                    creation_date__gte=date_ini,
                    state=True
                )
            # Documento(s)
            elif doc_values:
                consecutive_base = Answer_Consecutive.objects.filter(
                    form_consecutive__enterprise_id=user_val.enterprise_id,
                    form_consecutive__state=True,
                    form_consecutive_id__in=doc_values,
                    state=True
                )
            else:
                consecutive_base = Answer_Consecutive.objects.filter(
                    form_consecutive__enterprise_id=user_val.enterprise_id,
                    form_consecutive__state=True,
                    state=True
                )

            if (user_val.role_enterprise and not user_val.role_enterprise.view_all):
                answer_base = answer_base.filter(created_by=user_val)
                consecutive_base = consecutive_base.filter(created_by=user_val)

            if (public):
                answer_base = answer_base.filter(Q(created_by__in=usr_ids) | Q(created_by__isnull=public))
                consecutive_base = consecutive_base.filter(Q(created_by__in=usr_ids) | Q(created_by__isnull=public))
            elif (usr_ids):
                answer_base = answer_base.filter(Q(created_by__in=usr_ids))
                consecutive_base = consecutive_base.filter(Q(created_by__in=usr_ids))
            elif (role_val and role_val != '0'):
                answer_base = answer_base.filter(created_by__role_enterprise_id=role_val)
                consecutive_base = consecutive_base.filter(created_by__role_enterprise_id=role_val)
            

            answer_values = answer_base.select_related(
                'form_enterprise',
                'created_by'
            ).annotate(
                id_form=F('form_enterprise_id'),
                name_form=F('form_enterprise__name'),
                pin_form=F('form_enterprise__pin'),
            ).values(
                'id',
                'id_form',
                'name_form',
                'consecutive',
                'created_by__first_name',
                'created_by__first_last_name',
                'latitude',
                'longitude',
                'creation_date',
                'creation_date__date',
                'pin_form',
                'created_by_id'
            ).order_by('-creation_date')

            consecutive_values = consecutive_base.select_related(
                'form_consecutive',
                'created_by'
            ).annotate(
                consecutive=Value(True),
                id_form=F('form_consecutive_id'),
                name_form=F('form_consecutive__name'),
            ).values(
                'id',
                'id_form',
                'name_form',
                'consecutive',
                'created_by__first_name',
                'created_by__first_last_name',
                'latitude',
                'longitude',
                'creation_date',
                'creation_date__date',
                'created_by_id'
            ).order_by('-creation_date')

            data_values = chain(answer_values, consecutive_values)
            # Ordena
            _key_ = lambda k: (k['created_by_id'] if k['created_by_id'] else -1, k['creation_date__date'], k['id_form'])
            response_data = {}
            for k,g in itertools.groupby(sorted(data_values, key=_key_, reverse=True), key=_key_):
                temp_user = str(k[0])
                temp_date = str(k[1])
                temp_form = str(k[2])
                if temp_user not in response_data:
                    response_data[temp_user] = {}
                    response_data[temp_user]['count'] = 0
                if temp_date not in response_data[temp_user]:
                    response_data[temp_user][temp_date] = {}
                    response_data[temp_user][temp_date]['count'] = 0
                if temp_form not in response_data[temp_user][temp_date]:
                    temp_data = list(g)
                    temp_count = len(temp_data)
                    response_data[temp_user][temp_date][temp_form]= {
                        'data': temp_data,
                        'count': temp_count
                    }
                    response_data[temp_user][temp_date]['count'] += temp_count
                    response_data[temp_user]['count'] += temp_count

            response['status'] = True
            response['data'] = response_data
            status_response = status.HTTP_200_OK
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)


class FollowUserDetail(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            date = request.data['date']
            user = request.data['user']
            user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
            date_val = datetime.strptime(date, '%Y-%m-%d')
            date_val = date_val.replace(tzinfo=TZ_INFO)
            follow_user_vals = Follow_User.objects.filter(
                user__enterprise_id=user_val.enterprise_id,
                user_id=user,
                creation_date__date=date_val,
                state=True
            )
            
            offline_user_vals = Follow_User_Offline.objects.filter(
                user__enterprise_id=user_val.enterprise_id,
                user_id=user,
                creation_date__date=date_val,
                state=True
            )

            follow_user_vals = follow_user_vals.union(
                offline_user_vals
            ).order_by(
                'creation_date'
            ).values('latitude', 'longitude', 'creation_date')

            follow_user_vals = list(follow_user_vals)
            path, stops = smooth_followage(list(follow_user_vals))

            answer_logs = Traceability_User.objects.filter(
                user_id=user,
                creation_date__date=date_val,
                group__in=[53,61],
                action=1
            ).values('creation_date', 'description')

            response['status'] = True
            response['data'] = path
            response['stops'] = stops
            response['answers'] = list(answer_logs)
            status_response = status.HTTP_200_OK

        except KeyError:
            response['detail'] = "Faltan Parametros"
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)

import pandas as pd
import numpy as np
import math

def smooth_followage(data_init):
    # print(len(data_init))
    data = [
        {
            'geometry': {'x':float(point['latitude']), 'y':float(point['longitude'])},
            't': (point['creation_date'])
        } for point in smooth_trajectory(data_init)
    ]
    df_geo = pd.DataFrame.from_records(data).set_index('t')

    path_dict = df_geo.to_dict()
    path_stops = stop_detection(data_init, min_duration=timedelta(seconds=300), max_diameter=100)

    path = [{point.astimezone(TZ_INFO).strftime("%H:%M:%S"): [path_dict['geometry'][point]['x'], path_dict['geometry'][point]['y']] } for point in path_dict['geometry']]

    return path, path_stops

def stop_detection(data, min_duration=timedelta(seconds=60), max_diameter=1000):
    stops = []
    stop_starts = []
    stop_ends = []
    # 2.372873, -75.903950  2.373455, -75.903251
    for i, x_1 in enumerate(data):
        temp_index = None
        times = 0
        if len(stop_ends) and i < stop_ends[-1]:
            continue

        for j, x_2 in enumerate(data[i+1:]):
            index = i + 1 + j
            temp_point = x_1 if (temp_index) == None else data[temp_index]
            distance = geo_distance(temp_point, x_2)
            if distance <= max_diameter + (max_diameter * 0.05):
                temp_index = index
            else:
                times += 1
                if temp_index and ((data[temp_index]['creation_date'] - x_1['creation_date']) >= min_duration):
                    stops.append({
                        'geo': [x_1['latitude'], x_1['longitude']],
                        'start_time': x_1['creation_date'],
                        'end_time': data[temp_index]['creation_date'],
                        'duration': (data[temp_index]['creation_date'] - x_1['creation_date']).total_seconds()
                    })
                    stop_ends.append(index)
                    stop_starts.append(i)
                temp_index = None
                if(times > 2):
                    break
        if temp_index and (data[temp_index]['creation_date'] - x_1['creation_date']) >= min_duration:
            stops.append({
                'geo': [x_1['latitude'], x_1['longitude']],
                'start_time': x_1['creation_date'],
                'end_time': data[temp_index]['creation_date'],
                'duration': (data[temp_index]['creation_date'] - x_1['creation_date']).total_seconds()
            })
            stop_ends.append(temp_index)
            stop_starts.append(i)

    stops_final = []
    for i, stop in enumerate(stops):
        if i == 0:
            stops_final.append(stop)
        else:
            pre_geo = {'latitude': stops_final[-1]['geo'][0], 'longitude': stops_final[-1]['geo'][1]}
            actual_geo = {'latitude': stop['geo'][0], 'longitude': stop['geo'][1]}
            distance = geo_distance(pre_geo, actual_geo)
            delta_time = stop['start_time'] - stops_final[-1]['end_time']
            # print(distance, delta_time.total_seconds())
            if (distance <= max_diameter) and delta_time.total_seconds() > 0 and (delta_time <= min_duration):
                stops_final[-1]['end_time'] = stop['end_time']
                stops_final[-1]['duration'] = (stop['end_time'] - stops_final[-1]['start_time']).total_seconds()
            elif (distance <= max_diameter) and (delta_time.total_seconds() < 0):
                pass
            else:
                stops_final.append(stop)

    for i, stop in enumerate(stops_final):
        stops_final[i]['start_time'] = stop['start_time'].astimezone(TZ_INFO).strftime("%H:%M:%S")
        stops_final[i]['end_time'] = stop['end_time'].astimezone(TZ_INFO).strftime("%H:%M:%S")


    return stops_final

def geo_distance(initial, final):
    EARTHS_RADIO = 6371000
    delta_lat = math.radians(float(initial['latitude']) - float(final['latitude']))
    delta_lng = math.radians(float(initial['longitude']) - float(final['longitude']))

    temp = np.power(np.sin(delta_lat/2), 2) + np.cos(math.radians(float(initial['latitude']))) * \
        np.cos(math.radians(float(final['latitude']))) * np.power(np.sin(delta_lng/2), 2)

    c = 2 * np.arctan( np.power(temp, 0.5) / np.power(1-temp, 0.5) )
    return c * EARTHS_RADIO

def smooth_trajectory(data):
    result = [data[0]] if data else []
    for point in data[1:]:
        temp = result[-1]
        # Verificación de distancia menor a X metros
        if geo_distance(temp, point) > 120:
            result.append(point)
    return result
