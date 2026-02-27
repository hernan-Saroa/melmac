import itertools
from api.controllers.admin import get_enterprise
from api.controllers.api import check_usage_api_enterprise, register_api_usage
from api.controllers.routing import geocode_address
from api.models import User_Enterprise
from api.util import AddressNormalizer

from django.conf import settings
from django.http import Http404, HttpResponseBadRequest
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from threading import Thread
import json
import math
import openpyxl
import os
import pytz
import time

TZ_INFO = pytz.timezone('America/Bogota')

threads_list = {}

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def load_points(request, format=None):
    response = {"status": False}
    status_response = status.HTTP_400_BAD_REQUEST

    if "template" in request.FILES:
        try:
            enterprise = get_enterprise(request)
            name_file = str(request.FILES['template'])
            if name_file.lower().endswith(('.xlsx')):
                xlsx_file = request.FILES['template']
                wb_obj = openpyxl.load_workbook(xlsx_file)

                sheet = wb_obj.active

                # print('sheet.max_row, sheet.max_column')
                # print(sheet.max_row, sheet.max_column)

                if sheet.max_row > 1:

                    thread_ID = 'TH-'+str(enterprise)+'-'+str(len(threads_list)+1)
                    
                    new_thread = PointsLoadingThread(name_file, thread_ID, sheet,)
                    if str(enterprise) not in threads_list:
                        threads_list[str(enterprise)] = []
                        
                    threads_list[str(enterprise)].append(new_thread)
                    threads_list[str(enterprise)][-1].start()
                    
                    response['status'] = True
                    response['data'] = {
                        'thread_id': thread_ID,
                        'status': 0
                    }
                    status_response = status.HTTP_202_ACCEPTED
        except User_Enterprise.DoesNotExist:
            pass
        except Exception as e:
            response['detail'] = str(e)
    return Response(response, status_response)
        
class PointsLoadingThread(Thread):
    
    def __init__(self, thread_name, thread_ID, sheet):
        Thread.__init__(self)
        self.thread_name = thread_name
        self.thread_ID = thread_ID
        self.sheet = sheet
        self.daemon = True
        self.done = False
        self.json_data = None
        self.data_error = None
        self.total = 0
        self.processed = 0
 
        # helper function to execute the threads
    def run(self):
        print(str(self.thread_name) +" "+ str(self.thread_ID))
        self.run_route(self.sheet)
    
    def run_route(self, sheet):

        enterprise = self.thread_ID.split('-')[1]
        self.total = sheet.max_row - 1
        loc_file_name = '/dir.json'
        path = settings.MEDIA_ROOT

        if not os.path.exists(path):
            os.makedirs(path)
        path += loc_file_name
    
        try:
            content_file = ""
            with open(path, "r") as f:
                content_file = f.read()        
            loaded_addresses = json.loads(content_file)
        except:
            loaded_addresses = []
        list_addresses = { addr['address'] + ',' + addr['city']: addr for addr in loaded_addresses}
        data = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=3)):
            num_row = i + 2

            # columns data
            address = ''
            city = ''
            department = ''
            

            for cell in row:
                if cell.value is not None:
                    if cell.column == 1:
                        address = cell.value.strip()
                    elif cell.column == 2:
                        city = cell.value.strip()
                    elif cell.column == 3:
                        department = cell.value.strip()

            try:
                # print(address, city)
                if address == '' or city == '':
                    raise Exception('Datos incompletos')

                address_normalize = AddressNormalizer().normalize(address)
                if address_normalize:
                    if address_normalize[0]['address'] == "" and address_normalize[1]['address'] == "":
                        raise Exception('La dirección no pudo normalizarse')

                search_address = address_normalize[0]['address'] if address_normalize[0]['address'] != "" else address_normalize[1]['address']
                key = '{},{}'.format(address, city)
                if key not in list_addresses:    

                    if check_usage_api_enterprise(1, enterprise):
                        address_info = geocode_address(search_address, city)
                        # time.sleep(2)
                        register_api_usage(enterprise, 1, 500 if not address_info else 200)
                        if not address_info:
                            raise Exception('La dirección no fue encontrada')

                        data.append({
                            'address': address,
                            'address_normalize': address_normalize,
                            'city': city,
                            'department': department,
                            'lat': address_info['lat'],
                            'lon': address_info['lng'],
                            'zone': address_info,
                        })
                        loaded_addresses.append({
                            'address': address,
                            'address_normalize': address_normalize,
                            'city': city,
                            'department': department,
                            'lat': address_info['lat'],
                            'lon': address_info['lng'],
                            'zone': address_info,
                        })
                    
                    else:
                        raise Exception('Ya no tienes mas usos de la API, obten un nuevo paquete e intenta de nuevo.')
                else: 
                    data.append(list_addresses[key])
            except Exception as e:
                # print('error', num_row)
                if not self.data_error:
                    self.data_error = []
                self.data_error.append([str(e),
                    'Dirección: ' + address +
                    ', ciudad: ' + city + 
                    ', departamento: ' + department, num_row]
                )
            # print(num_row, "processed")
            self.processed += 1

        self.json_data = {
            'loaded': len(data), 
            'errors': len(self.data_error) if self.data_error else 0,
            'error': self.data_error
        }

        get_attr = lambda point: point['city'] if 'city' in point else 'N/A'
        points_list = {k: list(g) for k, g in itertools.groupby(sorted(data, key=get_attr), get_attr)}
        self.json_data['addresses'] = points_list
        
        with open(path, "w") as f:
            f.write(json.dumps(loaded_addresses))
        
        self.done = True
        # print('json_data__________:::')
        # print(self.json_data)
    

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def check_status_points(request):
    response = {
        'status': False
    }
    status_response = status.HTTP_400_BAD_REQUEST
    
    try:
        # print(request.data)
        if 'id' in request.data:
            thread_id = request.data['id']
            thread_selected = None
            enterprise = get_enterprise(request)
            index = -1
            if str(enterprise) in threads_list:
                for i, th in enumerate(threads_list[str(enterprise)]):
                    # print(th.thread_ID, thread_id)
                    if th.thread_ID == thread_id:
                        status_response = status.HTTP_200_OK
                        thread_selected = th
                        index = i
                        break
            
            if thread_selected is not None:
                # print(thread_selected.done)
                if thread_selected.done:
                    response['status'] = True
                    response['data'] = thread_selected.json_data
                    del threads_list[str(enterprise)][index]
                else:
                    response['data'] = {}
                    response['data']['total'] = thread_selected.total
                    response['data']['processed'] = thread_selected.processed
            else:
                response['detail'] = 'No se encuentra el proceso.'
                
    except Exception as e:
        print(e)
        response['status'] = False
        response['detail'] = 'No se encuentra el proceso.'

    return Response(response, status_response)
