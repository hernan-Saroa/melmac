# rest_framework
from datetime import datetime
from django.contrib.auth.models import User
from api.controllers.api import check_usage_api_enterprise
from rest_framework import mixins
from rest_framework import generics
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.views import APIView
# authentication
from api.permissions import IsUserAdminOrHasPermission
from rest_framework.permissions import IsAuthenticated
# models
from api.models import (
    User_Enterprise,
    Role_Enterprise,
    Massive_File,
    Massive_User,
    Massive_Errors,
    Notification_User,
)
from django.db.models import F, Q

# Serializer
from api.serializers import (
    MassiveFileSerializer,
    UserEnterpriseModelSerializer,
    UserUpdateModelSerializer
)
# Admin
from api.controllers.admin import get_enterprise
# Site
from api.controllers.site import create_user,restartToken
# utils
from api.encrypt import Encrypt
import openpyxl
from threading import Thread
from api.controllers.traceability import TZ_INFO, create_multiple_traceabilities, create_traceability
import re

class UserEnterpriseApiList(mixins.ListModelMixin,
                  mixins.CreateModelMixin,
                  generics.GenericAPIView):
    """User Enterprise api list view."""
    permission_classes = [IsUserAdminOrHasPermission]
    serializer_class = UserEnterpriseModelSerializer

    def get_queryset(self):
        # User active
        queryset = User_Enterprise.objects.filter(
            enterprise_id=self.kwargs.get('enterprise'),
            role_id=3,
            state=True
        )
        return queryset

    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        return Response(create_user(request.data))


class UserEnterpriseApiDetail(mixins.RetrieveModelMixin,
                    mixins.UpdateModelMixin,
                    generics.GenericAPIView):
    """User Enterprise api detail view."""
    permission_classes = [IsUserAdminOrHasPermission]
    serializer_class = UserUpdateModelSerializer

    def get_queryset(self):
        # User active
        queryset = User_Enterprise.objects.filter(
            enterprise_id=self.kwargs.get('enterprise'),
            role_id=3,
            state=True
        )
        return queryset

    def get(self, request, *args, **kwargs):
        return self.retrieve(request, *args, **kwargs)

    def put(self, request, *args, **kwargs):
        response = {}
        response['status'] = False
        data = request.data
        id = self.kwargs['pk']
        try:
            user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)

            user_enterprise_val = User_Enterprise.objects.get(id=id)
            previous_user = {
                'id': id,
                'first_name': user_enterprise_val.first_name,
                'middle_name': user_enterprise_val.middle_name,
                'first_last_name': user_enterprise_val.first_last_name,
                'second_last_name': user_enterprise_val.second_last_name,
                'login_state': user_enterprise_val.login_state,
                'login_count': user_enterprise_val.login_count,
                'phone': user_enterprise_val.phone,
                'role_enterprise_id': user_enterprise_val.role_enterprise_id,
                'identification': user_enterprise_val.identification,
                'email': user_enterprise_val.email,
            }
            user_enterprise_val.first_name = data['first_name'].strip()
            user_enterprise_val.middle_name = data['middle_name'].strip() if data['middle_name'] != None else ''
            user_enterprise_val.first_last_name = data['first_last_name'].strip()
            user_enterprise_val.second_last_name = data['second_last_name'].strip() if data['second_last_name'] != None else ''
            user_enterprise_val.login_state = data['login_state']
            user_enterprise_val.phone = data['phone']
            user_enterprise_val.identification = data['identification']
            rolUpdate = "0"
            if 'login_count' in data:
                user_enterprise_val.login_count = data['login_count']
            if data['role_enterprise_id'] != '':
                if str(data['role_enterprise_id']) != str(user_enterprise_val.role_enterprise_id):
                    user_enterprise_val.role_enterprise_id = data['role_enterprise_id']
                    rolUpdate = "1"
            else:
                user_enterprise_val.role_enterprise = None
            user_enterprise_val.save()

            docUser = User_Enterprise.objects.filter(enterprise_id=user_enterprise_val.enterprise_id, email=data['email']).exclude(id=id).count()
            if docUser == 0:
                if data['email'] != user_enterprise_val.email:
                    user_enterprise_val.email = data['email']
                    user_enterprise_val.save()
                    userF=restartToken(user_enterprise_val)
                    userF.username = data['email']
                    userF.email = data['email']
                    userF.save()
            if rolUpdate == "1":
                userF=restartToken(user_enterprise_val)


            changed_user = {
                'id': id,
                'first_name': user_enterprise_val.first_name,
                'middle_name': user_enterprise_val.middle_name,
                'first_last_name': user_enterprise_val.first_last_name,
                'second_last_name': user_enterprise_val.second_last_name,
                'login_state': user_enterprise_val.login_state,
                'login_count': user_enterprise_val.login_count,
                'role_enterprise_id': user_enterprise_val.role_enterprise_id,
                'phone': user_enterprise_val.phone,
                'identification': user_enterprise_val.identification,
                'email': user_enterprise_val.email,
            }

            check_changes_user(previous_user, changed_user, user_val)

            response['status'] = True
            if docUser == 0:
                response['optionResp'] = 1
                response['message'] = '¡Usuario actualizado correctamente!.'
            else:
                response['optionResp'] = 2
                response['message'] = 'Usuario actualizado sin el correo. Está ya se encuentra en la empresa.'
        except User_Enterprise.DoesNotExist:
            response['message'] = 'Error con el usuario'
        return Response(response)

    def delete(self, request, *args, **kwargs):
        response = {}
        response['status'] = False
        data = request.data
        id = self.kwargs['pk']
        try:
            user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)

            user_enterprise_val = User_Enterprise.objects.get(id=id)
            user_enterprise_val.state = False
            user_enterprise_val.save()
            response['status'] = True
            response['message'] = 'Usuario eliminado'

            log_content = {
                'user': user_val.id,
                'group': 25,
                'element': user_enterprise_val.id,
                'action': 2,
                'description': "El usuario denominado: #" + str(user_val.id) + " " + user_val.first_name + " " + user_val.first_last_name + " ha inabilitado al usuario #" + str(user_enterprise_val.id),
            }
            create_traceability(log_content)
        except User_Enterprise.DoesNotExist:
            response['message'] = 'Error con el usuario'
        return Response(response)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_role(request, enterprise):
    role_values = Role_Enterprise.objects.filter(enterprise_id=enterprise).values('id', 'name')
    return Response(role_values)


class MassiveUser(APIView):

    permission_classes = [IsUserAdminOrHasPermission]

    def get(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        try:
            enterprise = get_enterprise(request)
            mass_files = Massive_File.objects.filter(enterprise_id=enterprise, type=1).order_by('date')
            serializer = MassiveFileSerializer(mass_files, many=True)
            return Response({'status': True, 'data': serializer.data})
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)

    def post(self, request, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST

        if "template" in request.FILES:
            try:
                user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
                enterprise = user_val.enterprise_id
                name_file = str(request.FILES['template'])
                if name_file.lower().endswith(('.xlsx')):
                    xlsx_file = request.FILES['template']
                    wb_obj = openpyxl.load_workbook(xlsx_file)

                    sheet = wb_obj.active

                    print('sheet.max_row, sheet.max_column')
                    print(sheet.max_row, sheet.max_column)

                    if sheet.max_row > 1:
                        amount_rows = sheet.max_row - 1
                        massive_val = Massive_File()
                        massive_val.enterprise_id = enterprise
                        massive_val.amount = amount_rows
                        massive_val.type = 1
                        massive_val.template = request.FILES['template']
                        massive_val.save()
                        new_thread = Thread(target=run_user, args=(massive_val, sheet))
                        new_thread.start()

                        log_content = {
                            'user': user_val.id,
                            'group': 63,
                            'element': massive_val.id,
                            'action': 2,
                            'description': "El usuario denominado: #" + str(user_val.id) + " " + user_val.first_name + " " + user_val.first_last_name + " ha iniciado la carga masiva de usuarios #" + str(massive_val.id),
                        }
                        create_traceability(log_content)

                        response['status'] = True
                        response['data'] = {'id': massive_val.id, 'template': name_file, 'amount': amount_rows, 'success': 0, 'errors':0, 'progressbar':0, 'status':0}
                        status_response = status.HTTP_202_ACCEPTED

                        if not check_usage_api_enterprise(-1, enterprise, amount_rows):
                            response['warning'] = 'Puede que se registren parcialmente los usuarios debido al limite de usuarios para tu empresa.'
            except User_Enterprise.DoesNotExist:
                pass
            except Exception as e:
                print(e)
        return Response(response, status_response)


def run_user(massive, sheet):
    enterprise = massive.enterprise_id
    roles_values = Role_Enterprise.objects.filter(enterprise_id=massive.enterprise_id)
    roles = {category.name.upper() : category.id for category in list(roles_values)}

    type_ids = {
        'RC': 1, # 'Registro civil de nacimiento'
        'TI': 2, # 'Tarjeta de identidad'
        'CC': 3, # 'Cédula de ciudadanía'
        'CE': 5, # 'Cédula de extranjería'
    }

    log_list = []

    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=(massive.amount + 1), max_col=9)):
        num_row = i + 2
        email = ''
        first_name = ""
        middle_name = ""
        first_last_name = ""
        second_last_name = ""
        type_identification = None
        identification = ""
        phone = ''
        role_enterprise = None

        # Lectura de Excel
        for cell in row:
            if cell.value is not None:
                if cell.column == 1:
                    first_name = cell.value.strip()
                elif cell.column == 2:
                    middle_name = cell.value.strip()
                elif cell.column == 3:
                    first_last_name = cell.value.strip()
                elif cell.column == 4:
                    second_last_name = cell.value.strip()
                elif cell.column == 5:
                    type_identification = cell.value.strip().upper()
                elif cell.column == 6:
                    identification = str(cell.value).strip().upper()
                elif cell.column == 7:
                    email = cell.value.strip().lower()
                elif cell.column == 8:
                    phone = str(cell.value).strip()
                elif cell.column == 9:
                    role_enterprise = cell.value.strip().upper() if cell.value != "" else None

        # Verificación de datos incompletos
        try:
            error = ''
            if (first_name == '' or first_last_name == '' or type_identification == '' or
                identification == '' or email == '' or phone == ''):
                raise Exception('Datos incompletos')
            else:
                # Names
                tem_val = first_name.replace(' ', '')
                if not tem_val:
                    error += (", " if error != "" else "" ) + 'Primer Nombre'
                tem_val = first_last_name.replace(' ', '')
                if not tem_val:
                    error += (", " if error != "" else "" ) + 'Primer Apellido'

                # Identification
                tem_val = type_identification.replace(' ', '')
                if not tem_val:
                    error += (", " if error != "" else "" ) + 'Tipo Identificación'
                tem_val = identification.replace(' ', '')
                if not tem_val.isnumeric():
                    error += (", " if error != "" else "" ) + 'Identificación'

                # Correo
                tem_val = email.replace(' ', '')
                if not re.match("[^@ \t\r\n]+@[^@ \t\r\n]+\.[^@ \t\r\n]+", tem_val):
                    error += (", " if error != "" else "" ) + 'Correo Electronico'
                # Telefono
                tem_val = phone.replace(' ', '')
                if len(tem_val) != 10 or not tem_val.isnumeric():
                    error += (", " if error != "" else "" ) + 'Telefono'

                # Error
                if error != '':
                    error += (' invalidos' if ',' in error else ' invalido')
                    raise Exception(error)

            try:
                # Valida si existe el usuario
                user_val = User_Enterprise.objects.get(Q(email=email, enterprise_id=enterprise) | Q(phone=phone, enterprise_id=enterprise) | Q(identification=identification, enterprise_id=enterprise))
                raise Exception('El usuario ya esta registrado con correo, telefono o numero de identificación en esta empresa')
            except User_Enterprise.DoesNotExist:
                if not check_usage_api_enterprise(-1, enterprise):
                    raise Exception('Has superado el limite de usuarios para tu empresa, contacta a tu administrador para cambiar tu plan.')
                # Crear Usuario Django
                try:
                    user_new = User.objects.get(username=email)
                except User.DoesNotExist:
                    user_new = User()
                    user_new.username = email
                    user_new.first_name = first_name
                    user_new.last_name = first_last_name
                    user_new.email = email
                    user_new.save()

                # Crear Usuario Empresa
                user_val = User_Enterprise()
                user_val.enterprise_id = enterprise
                user_val.user = user_new
                user_val.first_name = first_name
                user_val.middle_name = middle_name
                user_val.first_last_name = first_last_name
                user_val.second_last_name = second_last_name
                user_val.identification = identification
                user_val.phone = phone
                user_val.email = email
                user_val.password = Encrypt().encrypt_code(identification)
                user_val.role_id = 3

                if type_identification in type_ids:
                    user_val.type_identification_id = type_ids[type_identification]
                else:
                    raise Exception('Dato Tipo Identificación no coincide con ninguna de las opciones ' + (", ".join(type_ids.keys())))

                if role_enterprise:
                    if role_enterprise in roles:
                        user_val.role_enterprise_id = roles[role_enterprise]
                    else:
                        raise Exception('Dato Rol no coincide con ninguna de las opciones ' + (", ".join(roles.keys())))

                user_val.save()

                massive_val = Massive_User()
                massive_val.massive_file = massive
                massive_val.user = user_val
                massive_val.save()

                log_content = {
                    'user': user_val.id,
                    'group': 25,
                    'element': user_val.id,
                    'action': 1,
                    'description': "El usuario denominado: #" + str(user_val.id) + " " + user_val.first_name + " " + user_val.first_last_name + " ha sido creado mediante la carga masiva de usuarios #" + str(massive.id),
                    'time': datetime.now(tz=TZ_INFO)
                }
                log_list.append(log_content)

                massive.success += 1
                massive.save()
        except Exception as e:
            massive.error += 1
            # role_enterprise = None
            fields = [
                ['primer nombre', first_name],
                ['segundo nombre', middle_name],
                ['primer apellido', first_last_name],
                ['segundo apellido', second_last_name],
                ['tipo id', type_identification],
                ['id', identification],
                ['correo', email],
                ['telefono', phone],
                ['rol empresa', role_enterprise]
            ]
            data_error = (str(e) + ' - ' +
                            (', '.join(
                                [str(field[0]) + ':' + str(field[1]) for field in fields]
                            )))
            error_val = Massive_Errors()
            error_val.massive_file = massive
            error_val.row = num_row
            error_val.data = data_error
            error_val.save()
    massive.status = 1
    massive.save()

    create_multiple_traceabilities(log_list)

class MassiveUserErrors(APIView):

    permission_classes = [IsUserAdminOrHasPermission]

    def get(self, request, pk, format=None):
        response = {"status": False}
        status_response = status.HTTP_400_BAD_REQUEST
        errors = []
        try:
            enterprise = get_enterprise(request)
            massive_errors_values = Massive_Errors.objects.filter(massive_file_id=pk, massive_file__type=1, massive_file__enterprise_id=enterprise)
            print(massive_errors_values.query)
            errors = []
            for massive_errors_val in massive_errors_values:
                error = {
                    'row': massive_errors_val.row,
                    'data': massive_errors_val.data
                }
                errors.append(error)
            response['status'] = True
            response['data'] = errors
            status_response = status.HTTP_200_OK
        except User_Enterprise.DoesNotExist:
            pass
        return Response(response, status_response)

def check_changes_user(previous:dict, changed:dict, user_val:User_Enterprise):
    parameters = ''
    if previous['first_name'] != changed['first_name']:
        parameters += 'Primer Nombre ' + previous['first_name'] + ' a ' + changed['first_name']

    if previous['middle_name'] != changed['middle_name']:
        parameters += 'Segundo Nombre ' + previous['middle_name'] + ' a ' + changed['middle_name']

    if previous['first_last_name'] != changed['first_last_name']:
        parameters += 'Primer Apellido ' + previous['first_last_name'] + ' a ' + changed['first_last_name']

    if previous['second_last_name'] != changed['second_last_name']:
        parameters += 'Segundo Apellido ' + previous['second_last_name'] + ' a ' + changed['second_last_name']

    if previous['login_state'] != changed['login_state']:
        parameters += 'Acceso ' + ('Activo' if previous['second_last_name'] else 'Desactivado') + ' a ' + ('Activo' if changed['second_last_name'] else 'Desactivado')

    if previous['role_enterprise_id'] != changed['role_enterprise_id']:
        parameters += 'Rol ' + str(previous['role_enterprise_id']) + ' a ' + str(changed['role_enterprise_id'])

    if previous['email'] != changed['email']:
        parameters += 'el correo ' + previous['email'] + ' a ' + changed['email']

    if str(previous['identification']) != str(changed['identification']):
            parameters += 'el número de identificación ' + str(previous['identification']) + ' por el número ' + str(changed['identification'])

    if previous['phone'] != changed['phone']:
        parameters += 'el teléfono ' + str(previous['phone']) + ' a ' + str(changed['phone'])


    if parameters != '':
        # Llamado del metodo de crear logs
        log_content = {
            'user': user_val.id,
            'group': 25,
            'element': previous['id'],
            'action': 2,
            'description': "El usuario denominado: #" + str(user_val.id) + " " + user_val.first_name + " " + user_val.first_last_name + " ha modificado " + parameters + " del usuario #" + str(previous['id']),
        }
        create_traceability(log_content)
        pass

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_notifications(request, pos):
    response = {"status": False}
    status_response = status.HTTP_400_BAD_REQUEST
    limit = 10
    try:
        user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
        notify_values = Notification_User.objects.filter(
            user=user_val
        ).exclude(
            status=2
        ).order_by(
            '-creation_date'
        ).annotate(
            date=F('creation_date')
        ).values(
            'title',
            'description',
            'type',
            'date',
            'id',
            'status',
            'extra',
        )
        response['status'] = True
        response['data'] = list(notify_values)[pos*limit: (pos+1)*limit]
        status_response = status.HTTP_200_OK
    except:
        pass
    return Response(response, status_response)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def set_notification_status(request):
    response = {"status": False}
    status_response = status.HTTP_400_BAD_REQUEST
    data = request.data
    try:
        user_val = User_Enterprise.objects.get(user=request.user, token=request.auth.key)
        Notification_User.objects.filter(
            user=user_val,
            id__in=data['id'],
        ).update(
            status=data['status'],
        )
        response['status'] = True
        status_response = status.HTTP_200_OK
    except:
        pass
    return Response(response, status_response)
